"""
Generate a wide-form FB cost-file Excel from the current Airtable state.

Output matches the layout of FB_Taverns_Cost_Price_File_Apr_26_v*.xlsx so the
file remains familiar to anyone who reads it. The export is read-only:
edits should happen in Airtable, then a fresh export can be regenerated.

Each cell is the price that BILLS on the as-of date (default today): one
winner per (site, product) by the shared precedence — an in-window support
first, then newest valid_from (reconcile.select_winners) — exactly what the
/master grid shows and the reconciler applies. A cell whose winner is a
temporary support is shaded and listed on the Info sheet with its window and
the standing price, so a re-upload of the file never silently bakes a support
price into the standard list.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from airtable_io import (  # noqa: E402
    _list_all,
    T,
    load_rules_from_airtable,
    load_sites_from_airtable,
)
from reconcile import is_support_rule, select_winners  # noqa: E402


def _billing_rules(rules, as_of: date | None) -> tuple[list, dict]:
    """(winners, standing): the rules that bill on ``as_of`` (today when None),
    one per (site, product) by the shared precedence, plus the standing
    (non-support) winner per key for annotating supported cells."""
    on = as_of or date.today()
    winners = list(select_winners(rules, on).values())
    standing = select_winners([r for r in rules if not is_support_rule(r)], on)
    return winners, standing


def _gather_state(as_of: date | None = None):
    rules = load_rules_from_airtable()
    sites = load_sites_from_airtable()
    products = _list_all(
        T["Products"],
        fields=["product_code", "description", "supplier", "retro_per_keg", "retro_eligible"],
    )
    products_by_code: dict[str, dict] = {}
    for rec in products:
        f = rec["fields"]
        code = f.get("product_code")
        if not code:
            continue
        products_by_code[code] = {
            "name": f.get("description", "") or "",
            "retro_per_keg": float(f.get("retro_per_keg") or 0.0),
            "supplier": f.get("supplier", "") or "",
        }

    active_rules, standing = _billing_rules(rules, as_of)
    return active_rules, sites, products_by_code, standing


def build_master_xlsx_bytes(as_of: date | None = None) -> bytes:
    """Fresh-sweep build (three full Airtable reads, ~30s at estate scale).
    NOT for request paths behind the hub proxy (~30s timeout) — the
    /export-master route uses build_master_xlsx_bytes_from_snapshot."""
    active_rules, sites, products_by_code, standing = _gather_state(as_of)
    return _build_xlsx(active_rules, sites, products_by_code, as_of, standing)


def build_master_xlsx_bytes_from_snapshot(snap, as_of: date | None = None) -> bytes:
    """Build from an already-loaded MasterSnapshot — sub-second, so the
    download works through the hub proxy. The snapshot is SWR-cached (≤60s
    stale at worst) and every grid edit re-publishes it patched, so the
    export always reflects the online master's current state."""
    active_rules, standing = _billing_rules(snap.rules, as_of)
    products_by_code = {
        code: {
            "name": (info.get("desc") or ""),
            "retro_per_keg": float(info.get("retro_per_keg") or 0.0),
            "supplier": "",
        }
        for code, info in (getattr(snap, "products", {}) or {}).items()
    }
    return _build_xlsx(active_rules, snap.sites, products_by_code, as_of, standing)


def _build_xlsx(active_rules, sites, products_by_code, as_of: date | None, standing=None) -> bytes:
    standing = standing or {}
    # site_ids: any site that currently has an active rule, sorted ascending
    site_ids = sorted({r.site_id for r in active_rules})
    site_name = lambda sid: (sites.get(sid) or {}).get("name", "") or sid
    site_account = lambda sid: (sites.get(sid) or {}).get("account_no", "") or ""

    # product_codes: anything with an active rule OR a retro on the master
    codes_active = {r.product_code for r in active_rules}
    codes_retro = {c for c, p in products_by_code.items() if p["retro_per_keg"] > 0}

    # fb_price per product: take from any billing rule for that product (they're
    # equal — the list price is product-level); standing rules first so a
    # support's copy is never the one picked.
    product_fb: dict[str, float] = {}
    product_name: dict[str, str] = {}
    for r in sorted(active_rules, key=is_support_rule):
        if r.product_code not in product_fb and r.fb_price:
            product_fb[r.product_code] = float(r.fb_price)
        if r.product_code not in product_name and r.product_desc:
            product_name[r.product_code] = r.product_desc
    for code, p in products_by_code.items():
        product_name.setdefault(code, p["name"])

    # Sort products alphabetically by name (case-insensitive); code as tiebreaker
    product_codes = sorted(
        codes_active | codes_retro,
        key=lambda c: (product_name.get(c, "").upper(), c),
    )

    # tenant prices: (site, product) -> price; supported cells noted separately
    tenant: dict[tuple[str, str], float] = {}
    supported: dict[tuple[str, str], object] = {}
    for r in active_rules:
        if r.tenant_price is not None:
            tenant[(r.site_id, r.product_code)] = float(r.tenant_price)
            if is_support_rule(r):
                supported[(r.site_id, r.product_code)] = r

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "FB Cost Price File"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDE6F2")
    support_fill = PatternFill("solid", fgColor="FFF4CC")
    border = Border(*([Side(style="thin", color="CCCCCC")] * 4))
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 1: each site's LWC account number (blank in the 5 left columns),
    #        matching the cost file whose top row holds the account codes.
    # Row 2: headers
    headers = ["Product Code", "Product Name", "Price", "Retro P/Keg", "Net price"]
    for sid in site_ids:
        headers.append(f"{site_name(sid)} {sid}")
    account_row = [None] * 5 + [site_account(sid) or None for sid in site_ids]
    ws.append(account_row)              # row 1 — account numbers
    ws.append(headers)                  # row 2
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = centre
        cell.border = border

    # Data rows
    for code in product_codes:
        fb = product_fb.get(code)
        retro = products_by_code.get(code, {}).get("retro_per_keg", 0.0) or 0.0
        net = fb - retro if fb is not None else None
        row = [code, product_name.get(code, ""), fb, retro if retro else None, net]
        for sid in site_ids:
            row.append(tenant.get((sid, code)))
        ws.append(row)
        for i, sid in enumerate(site_ids):
            if (sid, code) in supported:
                ws.cell(row=ws.max_row, column=6 + i).fill = support_fill

    # Number format: currency for cols 3..end; product code col stays as text
    for r in range(3, ws.max_row + 1):
        for c in range(3, ws.max_column + 1):
            ws.cell(row=r, column=c).number_format = '"£"#,##0.00'

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 38
    for c in range(3, ws.max_column + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 14

    # Freeze header + first two cols
    ws.freeze_panes = "C3"

    # Generated-on note in a second sheet so the wide format isn't disturbed
    info = wb.create_sheet("Info")
    info.append(["Generated", date.today().isoformat()])
    info.append(["As of", f"{as_of.isoformat()} (prices billing on that date)" if as_of
                 else "today (prices billing today)"])
    info.append(["Billing rules", len(active_rules)])
    info.append(["Sites covered", len(site_ids)])
    info.append(["Products listed", len(product_codes)])
    info.append([
        "Source of truth",
        "Airtable PricingRules + Products.retro_per_keg. Edit there, then re-export.",
    ])
    if supported:
        info.append([])
        info.append([
            "Supported prices",
            "Shaded cells are TEMPORARY tenant supports (they win for their window); "
            "the standing price resumes when the support ends. Do not re-upload them "
            "as standard prices.",
        ])
        for (sid, code), r in sorted(supported.items()):
            until = f"until {r.valid_to.isoformat()}" if r.valid_to else "open-ended"
            std = standing.get((sid, code))
            std_txt = (
                f"; standing price £{std.tenant_price:,.2f}"
                if std is not None and std.tenant_price is not None else ""
            )
            info.append([
                f"{sid} {site_name(sid)}",
                f"{code} {product_name.get(code, '')}: £{float(r.tenant_price):,.2f} {until}{std_txt}",
            ])
    for c in (1, 2):
        info.column_dimensions[get_column_letter(c)].width = 30

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
