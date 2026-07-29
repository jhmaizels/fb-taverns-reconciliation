"""
Seed the per-site tenant OFF-INVOICE layer (Site_Prices) into the Tennents
master workbook, from the team's existing individual price files.

Rates are NOT taken from the price files — those carry stale rates (Nick isn't
in the loop on the Tennents corrections; see [[tennents-pricing-position]]).
We take ONLY the per-site off-invoice discount (£/brl) — FB's internal split of
how much of the agreed discount the tenant gets on the invoice. Everything else
(WSP, agreed total discount, managed/bespoke construct) comes from the master
at export time, which auto-corrects the stale rates.

Output: a COPY of the master workbook with a new Site_Prices sheet + a bumped
README version. Review it, then upload it on /tennents (after the one-off
setup_tennents_tables.py migration that creates the TennentsSitePrices table).
NO Airtable writes here.

    python seed_tennents_site_prices.py \
        "FB_Taverns_Tennents_Master.xlsx" \
        "FB Taverns - Master Price File - July 2026.xlsx" \
        "FB_Taverns_Tennents_Master__with_prices.xlsx"
"""
from __future__ import annotations

import sys
import io
from datetime import date

import pandas as pd
from openpyxl import load_workbook

from tennents_master import parse_master_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# team price-file column positions (0-indexed), header at row 3/4, data row 6+
C_CODE, C_WSP, C_TOTAL, C_NETKEG, C_OFF = 1, 4, 5, 7, 9


def _num(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace("£", "").replace(",", ""))
    except ValueError:
        return None


def seed(master_path: str, pricefile_path: str, out_path: str) -> int:
    master = parse_master_workbook(master_path, "seed")
    site_by_name = {s.site_name.strip().upper(): s for s in master.sites}

    def find_sku(code: str):
        for part in str(code).replace("\\", "/").split("/"):
            part = part.strip()
            if not part:
                continue
            for cand in (part, part.zfill(6), part.lstrip("0")):
                s = master.find_sku(cand)
                if s:
                    return s
        return None

    book = pd.read_excel(pricefile_path, sheet_name=None, header=None)

    rows: dict[tuple, dict] = {}          # (account, canonical_sku) -> row dict
    flags = {"not_in_master": [], "no_agreed_rate": [], "rate_corrected": [],
             "keg_contradiction": [], "no_account": []}

    for sheet, df in book.items():
        site = sheet.strip()
        info = site_by_name.get(site.upper())
        if not info:
            flags["no_account"].append(site)
            continue
        acct = info.account
        for i in range(len(df)):
            code = df.iat[i, C_CODE]
            if code is None or (isinstance(code, float) and pd.isna(code)):
                continue
            code = str(code).strip()
            wsp = _num(df.iat[i, C_WSP])
            if not code or wsp is None:
                continue
            sku = find_sku(code)
            if not sku:
                flags["not_in_master"].append((site, code))
                continue
            from tennents_master import keg_brl_factor
            off = _num(df.iat[i, C_OFF]) or 0.0
            file_total = _num(df.iat[i, C_TOTAL])
            netkeg = _num(df.iat[i, C_NETKEG])

            # Pass an off-invoice through ONLY when the source keg positively
            # CONFIRMS it — i.e. the off implied by the source net/keg matches
            # the off column (per brl). Default is 0 (tenant on full WSP, all
            # retro): applying an unconfirmed off would cut the tenant's price
            # AND FB's retro margin (margin/brl = total − off). Every non-zero
            # off that is NOT confirmed — because the source keg is on full WSP,
            # is blank, or reflects a different value — stores 0 and is flagged
            # for the team to confirm and re-add deliberately.
            off_confirmed = 0.0
            if off > 0.01:
                off_from_keg = (wsp - netkeg / keg_brl_factor(sku)) if netkeg is not None else None
                if off_from_keg is not None and abs(off_from_keg - off) < 0.50:  # £/brl
                    off_confirmed = off
                else:
                    flags["keg_contradiction"].append((site, sku.sku_code, off))

            key = (acct, sku.sku_code)
            rows[key] = {
                "site": site, "account": acct, "sku": sku.sku_code,
                "product": sku.product or sku.brand, "off": round(off_confirmed, 2),
            }
            if sku.correct_total_per_brl is None:
                flags["no_agreed_rate"].append((site, sku.sku_code, sku.product))
            elif file_total is not None and abs(file_total - float(sku.correct_total_per_brl)) > 0.50:
                flags["rate_corrected"].append(
                    (site, sku.sku_code, file_total, float(sku.correct_total_per_brl)))

    # --- write Site_Prices into a copy of the master workbook ---
    wb = load_workbook(master_path)
    if "Site_Prices" in wb.sheetnames:
        del wb["Site_Prices"]
    ws = wb.create_sheet("Site_Prices")
    ws.append(["Site", "Tennents Account", "SKU Code", "Product",
               "Off-Invoice Discount £/Brl", "Notes"])
    for (acct, sku_code), r in sorted(rows.items(), key=lambda kv: (kv[1]["site"], kv[1]["sku"])):
        ws.append([r["site"], acct, sku_code, r["product"], r["off"], ""])

    # bump README version (README §5: any change bumps the version)
    if "README" in wb.sheetnames:
        rd = wb["README"]
        for row in rd.iter_rows():
            if row and str(row[0].value or "").lower().startswith("7. version"):
                cur = str(row[1].value or "")
                stamp = f" + Site_Prices seeded {date.today().isoformat()}"
                if "Site_Prices seeded" not in cur:
                    row[1].value = cur + stamp
                break

    wb.save(out_path)

    # --- report ---
    print(f"Seeded {len(rows)} (site, SKU) off-invoice rows across "
          f"{len({r['account'] for r in rows.values()})} sites → {out_path}\n")
    nonzero = sum(1 for r in rows.values() if r["off"] > 0.01)
    print(f"  off-invoice non-zero: {nonzero}   zero (full WSP, all retro): {len(rows) - nonzero}\n")

    def dump(title, items, fmt):
        uniq = sorted(set(items))
        print(f"  {title}: {len(uniq)}")
        for it in uniq[:60]:
            print("     " + fmt(it))
        print()

    print("=== REVIEW (confirm before/after upload) ===\n")
    dump("SKUs priced in files but NOT in master (add to master to price them)",
         [x[1] for x in flags["not_in_master"]], lambda c: f"code {c}")
    dump("Rate CORRECTED vs source file (net will change — this is intended)",
         flags["rate_corrected"],
         lambda t: f"{t[0]} / {t[1]}: file £{t[2]:.2f} → master £{t[3]:.2f}")
    dump("Priced in files but master has NO agreed rate (shown RATE TBC)",
         [(x[1], x[2]) for x in flags["no_agreed_rate"]], lambda t: f"{t[0]} {t[1]}")
    dump("Off-invoice entered in source but NOT applied there — stored as 0 (tenant "
         "stays on full WSP). Confirm any the tenant should get + we add them",
         flags["keg_contradiction"],
         lambda t: f"{t[0]} / {t[1]}: source off £{t[2]:.2f}/brl (not passed through)")
    if flags["no_account"]:
        dump("Price-file sheets with no master site", flags["no_account"], lambda s: s)
    return 0


def main(argv: list[str]) -> int:
    if len(argv) < 3:
        print(__doc__)
        return 2
    master_path = argv[1]
    pricefile_path = argv[2]
    out_path = argv[3] if len(argv) > 3 else "FB_Taverns_Tennents_Master__with_prices.xlsx"
    return seed(master_path, pricefile_path, out_path)


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
