"""
Generate the team-facing Tennents per-site price file from the master.

Nick Madigan's manual task (Jul-2026) was: for each site, copy the off-invoice
discount per brl into the "Pricer" tab, recompute the net keg price, build one
master workbook (a tab per site) and save an individual copy into each site's
folder. This module does all of that from the reconciliation master, so it
stays in sync: WSP + agreed total discount come from SKU_Master (a PINC flows
straight through) and the per-site off-invoice comes from the Site_Prices
layer (TennentsMaster.site_prices). See [[tennents-pricing-position]].

Three outputs, all built from a single TennentsMaster:
  build_master_workbook_bytes(master)   -> one .xlsx, a tab per site
  build_single_site_bytes(master, site) -> one .xlsx, that site only
  build_all_sites_zip_bytes(master)     -> a .zip of the single-site files

Per line the price shown is the AGREED (correct) rate from SKU_Master — NOT any
mis-loaded exception value — because the price file tells the tenant what they
SHOULD pay; open exceptions are surfaced in the Notes column instead.
"""
from __future__ import annotations

import io
import re
import zipfile
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.properties import PageSetupProperties

from tennents_master import (
    PINTS_PER_BRL,
    SiteInfo,
    SkuRate,
    TennentsMaster,
    keg_brl_factor,
    sku_codes,
)

# Column layout. Each site sheet stacks TWO purpose-built sections that share
# column widths: "Currently sold" (the site's live deals — current pricing) and
# "Substitution pricing" (products it doesn't sell, priced across a margin band).
# Cols A-E are common (identity + WSP + agreed total discount); Notes sits at the
# same far column in both so its width doesn't clash with a price column.
_ABV_COL = 3
_NOTES_COL = 13
_LASTCOL = _NOTES_COL

# "Currently sold" — what the tenant pays today. Editing the Tenant Off-Invoice
# (F) recalculates Net Keg/Pint (G/H) and FB Retro (I) live.
SOLD_HEADERS = {
    1: "SKU Code", 2: "Brand", 3: "ABV %", 4: "WSP £/Brl",
    5: "FB Taverns Total Discount", 6: "Tenant Off-Invoice £/Brl",
    7: "Net Price £/Keg", 8: "Net Price £/Pint", 9: "FB Retro £/Brl",
    _NOTES_COL: "Notes",
}
_SOLD_MONEY = {4, 5, 6, 7, 8, 9}

# "Substitution pricing" — for a product the site doesn't sell, at £150/£200/£250
# retained margin (RPB): the off-invoice to enter (Total Discount − RPB) AND the
# resulting tenant Net £/Keg, side by side.
SUB_HEADERS = {
    1: "SKU Code", 2: "Brand", 3: "ABV %", 4: "WSP £/Brl",
    5: "FB Taverns Total Discount", 6: "FB Net Cost £/Keg",
    7: "Off-Inv £/Brl @£150", 8: "Net £/Keg @£150",
    9: "Off-Inv £/Brl @£200", 10: "Net £/Keg @£200",
    11: "Off-Inv £/Brl @£250", 12: "Net £/Keg @£250",
    _NOTES_COL: "Notes",
}
_SUB_MONEY = {4, 5, 6, 7, 8, 9, 10, 11, 12}

_COL_WIDTHS = {
    1: 12, 2: 26, 3: 7, 4: 12, 5: 16,
    6: 15, 7: 14, 8: 14, 9: 15, 10: 14, 11: 15, 12: 14, _NOTES_COL: 44,
}

# Section order, matching the team's price file. Uncategorised SKUs fall into a
# trailing "Other" section (never hidden).
_CATEGORY_ORDER = ["Standard Lager", "Premium Lager", "Craft", "Ales & Stout", "Cider"]
_OTHER = "Other"

# Authoritative SKU -> section, taken from the team's file (correcting its
# off-by-one label placement: Caledonia Best is an ale, Drygate Pilsner is craft,
# Magners Ice Cold / Outcider are ciders). Keyed by primary AND alt codes so a
# lookup works whichever the master stored. Blackthorn (401176/401175) is mapped
# ahead of being added to the master.
_SKU_CATEGORY = {
    # Standard Lager
    "090425": "Standard Lager", "400889": "Standard Lager", "T00045238": "Standard Lager",
    # Premium Lager
    "401136": "Premium Lager", "400217": "Premium Lager", "400751": "Premium Lager",
    "400557": "Premium Lager", "401211": "Premium Lager", "STE025": "Premium Lager",
    "T00048477": "Premium Lager", "SAN002": "Premium Lager", "EST002": "Premium Lager",
    "T00043388": "Premium Lager", "KIN003": "Premium Lager",
    # Craft
    "401080": "Craft", "401079": "Craft", "DRY025": "Craft", "DIS009": "Craft",
    "T00044490": "Craft", "INN008": "Craft", "400248": "Craft", "INN005": "Craft",
    "JUB002": "Craft", "JUB003": "Craft",
    # Ales & Stout
    "400076": "Ales & Stout", "400745": "Ales & Stout", "004723": "Ales & Stout",
    "401152": "Ales & Stout", "004790": "Ales & Stout", "GUI002": "Ales & Stout",
    "GUI003": "Ales & Stout", "MCE005": "Ales & Stout", "MCE015": "Ales & Stout",
    "MCE008": "Ales & Stout", "T00045771": "Ales & Stout",
    # Cider
    "401172": "Cider", "401219": "Cider", "401173": "Cider", "401224": "Cider",
    "401178": "Cider", "401223": "Cider", "401174": "Cider", "401222": "Cider",
    "401188": "Cider", "401218": "Cider", "401176": "Cider", "401175": "Cider",
}

# Best-effort fallback for a SKU not in the map above (a future addition), so it
# still lands in a sensible section rather than "Other" where possible.
_CATEGORY_KEYWORDS = [
    ("Cider",         ("cider", "magners", "blackthorn", "outcider", "gaymers", "orchard", "olde english")),
    ("Ales & Stout",  ("stout", "guinness", "mcewan", "80/-", "70/-", "60/-", "bitter", "smooth", " ale")),
    ("Craft",         ("ipa", "pale", "drygate", "gladeye", "innis", "i&g", "jubel", "craft")),
    ("Premium Lager", ("heverlee", "menabrea", "bavarian", "stella", "mahou", "san miguel", "estrella", "kingfisher")),
    ("Standard Lager", ("lager", "pilsner")),
]


def _category(sku: SkuRate) -> str:
    for u in sku_codes(sku):          # primary + EVERY alt (alt_code may be '/'-joined)
        for cand in (u, u.zfill(6), u.lstrip("0")):
            if cand in _SKU_CATEGORY:
                return _SKU_CATEGORY[cand]
    hay = f"{sku.brand} {sku.product}".lower()
    for label, keys in _CATEGORY_KEYWORDS:
        if any(k in hay for k in keys):
            return label
    return _OTHER


def _sanitize_tab(name: str, used: set[str]) -> str:
    """Excel tab name: ≤31 chars, none of []:*?/\\, unique."""
    clean = re.sub(r"[\[\]:*?/\\]", " ", name).strip()[:31] or "Site"
    base = clean
    i = 2
    while clean.upper() in used:
        suffix = f" {i}"
        clean = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(clean.upper())
    return clean


def _price_line(master: TennentsMaster, site: SiteInfo, sku: SkuRate) -> dict:
    """Compute the price row for (site, sku). off-invoice/retro split follows the
    site construct: managed = all off-invoice (zero retro); bespoke flat retro =
    fixed retro; standard = the stored per-site off-invoice. Rate is the AGREED
    correct total from SKU_Master."""
    total = sku.correct_total_per_brl
    wsp = sku.wsp_per_brl
    note_bits: list[str] = []
    # Is there a live off-invoice DEAL for this (site, SKU)? A stored off-invoice
    # above £0 = a negotiated discount in place = the product is being sold here
    # (James, Jul-2026). Row-presence alone isn't it — the seed defaults unconfirmed
    # lines to £0, and a £0 line is full WSP, not a deal. Managed sites aren't split.
    has_deal = master.off_invoice(site.account, sku.sku_code) > 0.0

    if total is None:
        note_bits.append("RATE TBC — no agreed Tennents rate")
        return {
            "category": _category(sku), "code": sku.sku_code, "brand": sku.product or sku.brand,
            "abv": sku.abv, "wsp": wsp, "total": None, "net_brl": None,
            "net_keg": None, "net_pint": None, "off": None, "retro": None, "bpu": None,
            "has_deal": has_deal,
            "note": "; ".join(note_bits),
        }

    total = float(total)
    # off-invoice / retro split. Managed sites take the whole discount
    # off-invoice, zero retro (README §4 — a documented rule). Every other site
    # (incl. bespoke, whose split the master deliberately does NOT model — "flat
    # £200/brl retro: validate the total, not the split") uses the stored
    # per-site off-invoice, defaulting to 0 (tenant pays full WSP, all retro).
    if site.is_managed:
        off = total
    else:
        off = master.off_invoice(site.account, sku.sku_code)
    off = min(max(off, 0.0), total)      # never negative retro / over-100% off
    retro = total - off

    bpu = keg_brl_factor(sku)
    invoice_net_brl = (wsp - off) if wsp is not None else None
    row = {
        "category": _category(sku), "code": sku.sku_code, "brand": sku.product or sku.brand,
        "abv": sku.abv, "wsp": wsp, "total": total,
        "net_brl": (wsp - total) if wsp is not None else None,
        "net_keg": (invoice_net_brl * bpu) if invoice_net_brl is not None else None,
        "net_pint": (invoice_net_brl / PINTS_PER_BRL) if invoice_net_brl is not None else None,
        "off": off, "retro": retro, "bpu": bpu,
        "has_deal": has_deal,
        "note": "",
    }

    ex = None
    for code in sku_codes(sku):       # exceptions key on the RAW code — try each alt too
        ex = master._exception_index.get((site.account, code))
        if ex is not None:
            break
    if ex is not None and ex.loaded_total_per_brl is not None:
        note_bits.append(f"Correction pending — currently loaded £{ex.loaded_total_per_brl:,.2f}/brl")
    if site.is_managed:
        note_bits.append("Managed: full discount off-invoice")
    elif site.is_bespoke:
        note_bits.append("Bespoke construct — retro per agreement; total validated")
    row["note"] = "; ".join(note_bits)
    return row


def _by_category(lines: list[dict]) -> list[tuple[str, list[dict]]]:
    """Group rows into beer-type sections, canonical order (Other trailing)."""
    grouped: dict[str, list[dict]] = {}
    for ln in lines:
        grouped.setdefault(ln["category"] or _OTHER, []).append(ln)
    order = [c for c in _CATEGORY_ORDER if c in grouped]
    order += [c for c in grouped if c not in _CATEGORY_ORDER]
    return [(c, grouped[c]) for c in order]


def _row_cells(r: int, line: dict, kind: str) -> tuple[dict, dict]:
    """(values, formulas) keyed by 1-indexed column for one product row. kind
    'sold' = current-deal columns (editing Off-Invoice F drives Net Keg/Pint G/H
    and FB Retro I live); 'sub' = substitution pricing (FB cost + the off-invoice
    to enter AND the resulting tenant net keg at each £150/£200/£250 RPB). Cols
    A-E (identity, WSP, total) and Notes are common to both."""
    vals = {1: line["code"], 2: line["brand"], 3: line["abv"],
            4: line["wsp"], 5: line["total"], _NOTES_COL: line["note"]}
    forms: dict[int, str] = {}
    bpu = line["bpu"]
    priceable = line["total"] is not None and line["net_keg"] is not None
    if kind == "sold":
        vals[6] = line["off"]                                    # F  Tenant Off-Invoice (input)
        if line["total"] is not None:
            forms[9] = f"=E{r}-F{r}"                             # I  FB Retro = total - off
            if priceable:                                       # WSP present
                forms[7] = f"=(D{r}-F{r})*{bpu!r}"             # G  Net £/Keg  = (WSP - off) * keg factor
                forms[8] = f"=(D{r}-F{r})/{PINTS_PER_BRL!r}"   # H  Net £/Pint = (WSP - off) / pints per brl
    elif priceable:                                             # kind == "sub"
        forms[6]  = f"=(D{r}-E{r})*{bpu!r}"                     # F  FB net cost £/keg = (WSP - total) * keg
        forms[7]  = f"=E{r}-150"                                # G  off-invoice to enter for a £150 retro
        forms[8]  = f"=(D{r}-E{r}+150)*{bpu!r}"                # H  resulting tenant net £/keg
        forms[9]  = f"=E{r}-200"
        forms[10] = f"=(D{r}-E{r}+200)*{bpu!r}"
        forms[11] = f"=E{r}-250"
        forms[12] = f"=(D{r}-E{r}+250)*{bpu!r}"
    return vals, forms


def _write_site_sheet(ws, master: TennentsMaster, site: SiteInfo, as_of: date) -> None:
    bold = Font(bold=True)
    title_font = Font(bold=True, size=13, color="1F3B57")
    sub_font = Font(size=9, color="555555")
    header_fill = PatternFill("solid", fgColor="DDE6F2")
    tbc_fill = PatternFill("solid", fgColor="FBE9D6")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["A1"] = "FB Taverns — Tennents Scotland Price File"
    ws["A1"].font = title_font
    ws["A2"] = f"Site: {site.site_name}    |    Tennents account: {site.account}    |    Prices effective: {as_of.isoformat()}"
    ws["A2"].font = sub_font
    ws["A3"] = f"Operating model: {site.operating_model or '—'}    |    Discount construct: {site.discount_construct or '—'}"
    ws["A3"].font = sub_font

    section_fill = PatternFill("solid", fgColor="1F3B57")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    cat_fill = PatternFill("solid", fgColor="4A6A8A")
    cat_font = Font(bold=True, size=10, color="FFFFFF")

    def _band(row, text, fill, font):
        ws.cell(row=row, column=1, value=text).font = font
        for c in range(1, _LASTCOL + 1):
            ws.cell(row=row, column=c).fill = fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_LASTCOL)

    def _section(row, title, headers, money, rows, kind):
        _band(row, title, section_fill, section_font)
        row += 1
        for c in headers:                              # header cells (this section's columns)
            hc = ws.cell(row=row, column=c, value=headers[c])
            hc.font = bold
            hc.fill = header_fill
            hc.alignment = centre
            hc.border = border
        row += 1
        if not rows:
            ws.cell(row=row, column=1, value="— none —").font = Font(italic=True, color="888888")
            return row + 1
        for si, (cat, cat_rows) in enumerate(_by_category(rows)):
            if si:
                row += 1                               # blank spacer between beer types
            _band(row, cat, cat_fill, cat_font)
            row += 1
            for line in cat_rows:
                is_tbc = line["total"] is None
                vals, forms = _row_cells(row, line, kind)
                for c in headers:
                    v = vals.get(c)
                    cell = ws.cell(row=row, column=c, value=forms.get(c, v))
                    cell.border = border
                    if c in money and (isinstance(v, (int, float)) or c in forms):
                        cell.number_format = '"£"#,##0.00'
                    if c == _ABV_COL and isinstance(v, (int, float)):
                        cell.number_format = '0.0"%"'
                    if is_tbc:
                        cell.fill = tbc_fill
                row += 1
        return row

    lines = [_price_line(master, site, sku) for sku in master.skus]

    r = 5
    if site.is_managed:
        # Managed: FB runs the site — whole discount is off-invoice, no tenant
        # "deal vs not" split. One current-pricing list of everything.
        r = _section(r, "All products — managed site (whole discount off-invoice)",
                     SOLD_HEADERS, _SOLD_MONEY, lines, "sold")
    else:
        sold = [ln for ln in lines if ln["has_deal"]]
        rest = [ln for ln in lines if not ln["has_deal"]]
        r = _section(r, "Currently sold at this site — off-invoice deal in place",
                     SOLD_HEADERS, _SOLD_MONEY, sold, "sold")
        r += 1                                         # spacer, then a page break
        ws.row_breaks.append(Break(id=r - 1))          # substitution list prints as page 2
        r = _section(r, "Substitution pricing — products not currently sold (£150 / £200 / £250 RPB)",
                     SUB_HEADERS, _SUB_MONEY, rest, "sub")

    warn = ("" if master.site_prices else
            "WARNING: the per-site off-invoice layer (Site_Prices) is NOT loaded — every product shows "
            "as NOT sold and at full WSP. Seed/upload Site_Prices before using these figures. ")
    foot = ws.cell(
        row=r + 1, column=1,
        value=(warn + f"Generated {as_of.isoformat()} from the Tennents master "
               f"({master.version or 'version n/a'}). WSP & agreed total discount from SKU_Master; "
               f"per-site off-invoice from Site_Prices. TWO sections — 'Currently sold' = products with "
               f"an off-invoice deal in place (edit the Tenant Off-Invoice and Net Keg/Pint & FB Retro "
               f"recalculate); 'Substitution pricing' = products not currently sold, priced at £150 / "
               f"£200 / £250 retained margin (RPB), showing the off-invoice to enter (Total Discount − "
               f"RPB) and the resulting tenant Net £/Keg side by side. When switching a product the "
               f"replacement's margin must beat the retired product's — every switch accretive. A "
               f"negative off-invoice = the RPB is above the product's total discount (a price above "
               f"WSP). RATE TBC = no agreed Tennents rate yet — not priced."))
    foot.font = Font(size=9, bold=bool(warn), color="B00020" if warn else "555555")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=_LASTCOL)

    for c, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A7"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


def _sites_in_scope(master: TennentsMaster) -> list[SiteInfo]:
    """Estate sites with a real account, alphabetical."""
    seen: set[str] = set()
    out: list[SiteInfo] = []
    for s in master.sites:
        if not s.account or s.account.upper() == "TBC" or s.account in seen:
            continue
        seen.add(s.account)
        out.append(s)
    return sorted(out, key=lambda s: s.site_name.upper())


def list_site_choices(master: TennentsMaster) -> list[tuple[str, str]]:
    """(account, site_name) for the per-site download picker."""
    return [(s.account, s.site_name) for s in _sites_in_scope(master)]


def find_site(master: TennentsMaster, account_or_name: str) -> SiteInfo | None:
    key = str(account_or_name).strip().upper()
    for s in _sites_in_scope(master):
        if s.account.upper() == key or s.site_name.strip().upper() == key:
            return s
    return None


def build_master_workbook_bytes(master: TennentsMaster, as_of: date | None = None) -> bytes:
    """One workbook, a tab per site (the 'master which has everything')."""
    as_of = as_of or date.today()
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for site in _sites_in_scope(master):
        ws = wb.create_sheet(_sanitize_tab(site.site_name, used))
        _write_site_sheet(ws, master, site, as_of)
    if not wb.sheetnames:                          # never leave an empty workbook
        wb.create_sheet("No sites")
    return _to_bytes(wb)


def build_single_site_bytes(master: TennentsMaster, site: SiteInfo, as_of: date | None = None) -> bytes:
    """One workbook, one site — the individual copy for a site folder."""
    as_of = as_of or date.today()
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(_sanitize_tab(site.site_name, set()))
    _write_site_sheet(ws, master, site, as_of)
    return _to_bytes(wb)


def build_all_sites_zip_bytes(master: TennentsMaster, as_of: date | None = None) -> bytes:
    """A .zip of one single-site workbook per site, named for the site — ready
    to drop into each site's 'Price Lists and Bar Plans' folder."""
    as_of = as_of or date.today()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for site in _sites_in_scope(master):
            safe = re.sub(r'[\\/:*?"<>|]', " ", site.site_name).strip()
            fname = f"{safe} - Tennents Price File {as_of.isoformat()}.xlsx"
            zf.writestr(fname, build_single_site_bytes(master, site, as_of))
    return buf.getvalue()


def _to_bytes(wb: Workbook) -> bytes:
    # openpyxl writes formulas without a cached result, so tell Excel to
    # recalculate on open — otherwise the live Net/Retro cells read blank/0 until
    # the sheet is touched.
    wb.calculation.fullCalcOnLoad = True
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
