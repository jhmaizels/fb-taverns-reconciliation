"""
Offline tests for the master-editor foundation (design §5.2, phases 0-1):

  - ``airtable_io.end_pricing_rule``  (the new targeted-close primitive)
  - ``master_changes.validate_master_change``  (pure, table-driven)
  - ``master_changes.preview_master_change``   (pure: winner-on-date, warnings)
  - ``master_changes.apply_master_change``     (dispatch onto the primitives,
    including THROUGH-THE-EDITOR re-assertions of the upsert close guard)

No Airtable network access: reuses the projection-honouring fake
``_list_all`` / recording ``_batch`` pattern (and the synthetic master
fixtures) from test_upsert_close_guard.py.

Phase 2 adds the /master* route smoke tests (FastAPI TestClient, auth
stubbed at auth_supabase._resolve_supabase, Airtable via the same fakes,
and the async cache-refresh kick disarmed so no background thread can
outlive the fakes and attempt a real network call).

Run standalone (exit 0 = pass, 1 = fail):

    python test_master_editor.py

Also discoverable by pytest, but pytest is not a project dependency.
"""

from __future__ import annotations

import os
import re
from dataclasses import replace
from datetime import date, timedelta
from html import unescape

os.environ.setdefault("AIRTABLE_TOKEN", "test-token")
os.environ.setdefault("AIRTABLE_BASE_ID", "appTEST")

import airtable_io  # noqa: E402
from airtable_io import MasterSnapshot  # noqa: E402
from master_changes import (  # noqa: E402
    MasterChange,
    apply_master_change,
    change_rule_key,
    preview_master_change,
    validate_master_change,
)
from reconcile import Rule  # noqa: E402

# Reuse the close-guard test's synthetic master (same site/product fixtures).
from test_upsert_close_guard import (  # noqa: E402
    PROD_CODE,
    PROD_REC,
    PRODUCTS,
    SITE_ID,
    SITE_REC,
    SITES,
    _pr_record,
)

SITE2_ID, SITE2_REC = "002", "rec_site_002"
PR = lambda: airtable_io.T["PricingRules"]  # noqa: E731


def _key(vf: str | None) -> str:
    return f"{SITE_ID}|{PROD_CODE}|{vf or 'open'}"


class FakeAirtable:
    """In-memory Airtable: projection-honouring _list_all + recording _batch
    (the exact contract of test_upsert_close_guard's fakes), plus an
    invalidate_master_cache counter. Context manager installs/restores."""

    def __init__(self, pricing, sites=None, products=None):
        self.tables = {
            airtable_io.T["Sites"]: [dict(r) for r in (sites if sites is not None else SITES)],
            airtable_io.T["Products"]: [dict(r) for r in (products if products is not None else PRODUCTS)],
            airtable_io.T["PricingRules"]: [dict(r) for r in pricing],
        }
        self.calls: list[tuple] = []
        self.invalidations = 0

    def _list_all(self, table_id, fields=None, filter_by_formula=None):
        # filter_by_formula is a server-side perf scope (verified against live
        # Airtable separately); the fake returns the full set and lets the
        # in-Python close pass do the filtering the tests actually assert on.
        out = []
        for rec in self.tables.get(table_id, []):
            f = rec["fields"]
            if fields is not None:
                # Mimic Airtable: only the projected fields come back.
                f = {k: f[k] for k in fields if k in f}
            out.append({"id": rec["id"], "fields": dict(f)})
        return out

    def _batch(self, records, op, table_id):
        self.calls.append((op, table_id, [dict(r) for r in records]))
        if op == "create":
            created = []
            for i, r in enumerate(records):
                rec = {
                    "id": f"rec_new_{len(self.tables[table_id]) + i}",
                    "fields": dict(r["fields"]),
                }
                # Persist so follow-up lookups (e.g. auto-created sites) resolve.
                self.tables[table_id].append(rec)
                created.append(rec)
            return created
        return [{"id": r.get("id"), "fields": r.get("fields", {})} for r in records]

    def __enter__(self):
        self._orig = (
            airtable_io._list_all,
            airtable_io._batch,
            airtable_io.invalidate_master_cache,
        )
        airtable_io._list_all = self._list_all
        airtable_io._batch = self._batch

        def _inval():
            self.invalidations += 1

        airtable_io.invalidate_master_cache = _inval
        return self

    def __exit__(self, *exc):
        (
            airtable_io._list_all,
            airtable_io._batch,
            airtable_io.invalidate_master_cache,
        ) = self._orig
        return False


def _updates(fa: FakeAirtable) -> list[dict]:
    return [r for op, t, recs in fa.calls if op == "update" and t == PR() for r in recs]


def _creates(fa: FakeAirtable, table_id=None) -> list[dict]:
    tid = table_id or PR()
    return [r for op, t, recs in fa.calls if op == "create" and t == tid for r in recs]


def _rule(
    vf=None, vt=None, tenant=100.0, fb=120.0, retro=0.0,
    status="tenanted", site=SITE_ID, code=PROD_CODE, reason="orig",
) -> Rule:
    return Rule(
        site_id=site, product_code=code, product_desc="Test Keg",
        tenant_price=tenant, fb_price=fb, retro_pct=retro,
        valid_from=vf, valid_to=vt, status=status, reason=reason, source="test",
    )


def _snap(rules: list[Rule]) -> MasterSnapshot:
    """Pure MasterSnapshot over the standard synthetic site/product."""
    rule_ids = {
        airtable_io._rule_key(r.site_id, r.product_code, r.valid_from): f"rec_{i}"
        for i, r in enumerate(rules)
    }
    return MasterSnapshot(
        sites={SITE_ID: {"name": "Test Tavern", "status": "tenanted",
                         "country": "england", "notes": "", "_rec_id": SITE_REC}},
        rules=rules,
        site_ids={SITE_ID: SITE_REC},
        product_ids={PROD_CODE: PROD_REC},
        rule_ids=rule_ids,
        banner_info={},
    )


# --------------------------------------------------------------------------
# end_pricing_rule
# --------------------------------------------------------------------------

def test_end_rule_closes_open_rule():
    rec = _pr_record("rec_open", _key("2026-01-01"), valid_from="2026-01-01")
    rec["fields"]["reason"] = "original reason"
    with FakeAirtable([rec]) as fa:
        rec_id = airtable_io.end_pricing_rule(
            _key("2026-01-01"), date(2026, 7, 1), "delisted", "editor:me@x"
        )
    assert rec_id == "rec_open", f"expected rec_open, got {rec_id}"
    ups = _updates(fa)
    assert len(ups) == 1, f"expected exactly one PATCH, got {ups}"
    f = ups[0]["fields"]
    assert set(f) == {"valid_to", "reason", "source"}, (
        f"PATCH must carry valid_to+reason+source and NOTHING else, got {sorted(f)}"
    )
    assert f["valid_to"] == "2026-07-01"
    assert f["reason"] == "delisted; original reason", (
        f"reason must be prepended, not replaced: {f['reason']!r}"
    )
    assert f["source"] == "editor:me@x"
    assert fa.invalidations == 1, (
        f"end_pricing_rule must invalidate the master cache exactly once, got {fa.invalidations}"
    )


def test_end_rule_no_old_reason():
    rec = _pr_record("rec_open", _key("2026-01-01"), valid_from="2026-01-01")
    with FakeAirtable([rec]) as fa:
        airtable_io.end_pricing_rule(_key("2026-01-01"), date(2026, 7, 1), "delisted", "editor:me@x")
    assert _updates(fa)[0]["fields"]["reason"] == "delisted"


def test_end_rule_missing_key_raises():
    with FakeAirtable([]) as fa:
        try:
            airtable_io.end_pricing_rule(_key("2026-01-01"), date(2026, 7, 1), "r", "s")
        except ValueError as e:
            assert "no pricing rule" in str(e)
        else:
            raise AssertionError("missing rule_key must raise ValueError")
    assert not fa.calls or all(op != "update" for op, *_ in fa.calls), "nothing may be written"
    assert fa.invalidations == 0, "no cache invalidation on refusal"


def test_end_rule_already_ended_raises():
    rec = _pr_record("rec_done", _key("2026-01-01"), valid_from="2026-01-01", valid_to="2026-06-01")
    with FakeAirtable([rec]) as fa:
        try:
            airtable_io.end_pricing_rule(_key("2026-01-01"), date(2026, 7, 1), "r", "s")
        except ValueError as e:
            assert "already ended" in str(e)
        else:
            raise AssertionError("already-ended rule must raise ValueError")
    assert fa.invalidations == 0


def test_end_rule_inverted_interval_raises():
    rec = _pr_record("rec_open", _key("2026-06-01"), valid_from="2026-06-01")
    with FakeAirtable([rec]) as fa:
        for bad in (date(2026, 6, 1), date(2026, 1, 1)):  # equal and earlier
            try:
                airtable_io.end_pricing_rule(_key("2026-06-01"), bad, "r", "s")
            except ValueError as e:
                assert "must be after" in str(e)
            else:
                raise AssertionError(f"valid_to={bad} <= valid_from must raise ValueError")
    assert not _updates(fa) and fa.invalidations == 0


# --------------------------------------------------------------------------
# apply_master_change: price_change (incl. through-the-editor close-guard)
# --------------------------------------------------------------------------

def _price_change(d: date, tenant=182.0) -> MasterChange:
    return MasterChange(
        op="price_change", site_id=SITE_ID, product_code=PROD_CODE,
        tenant_price=tenant, valid_from=d, reason="LWC list increase",
    )


def test_price_change_closes_prior_open_and_creates_successor():
    sites = SITES + [{"id": SITE2_REC, "fields": {"site_id": SITE2_ID}}]
    other_site_rec = {
        "id": "rec_other_site",
        "fields": {
            "rule_key": f"{SITE2_ID}|{PROD_CODE}|2026-01-01",
            "valid_from": "2026-01-01",
            "site": [SITE2_REC], "product": [PROD_REC],
        },
    }
    existing = [
        _pr_record("rec_old", _key("2026-01-01"), valid_from="2026-01-01"),
        other_site_rec,
    ]
    with FakeAirtable(existing, sites=sites) as fa:
        res = apply_master_change(_price_change(date(2026, 7, 2)), "me@x")
    assert (res.created, res.updated, res.closed) == (1, 0, 1), (
        f"expected (1,0,1), got {(res.created, res.updated, res.closed)}"
    )
    closes = [u for u in _updates(fa) if set(u["fields"]) == {"valid_to"}]
    assert [u["id"] for u in closes] == ["rec_old"], (
        f"only THIS site's prior open rule may be closed (keys_in_new scoping), got {closes}"
    )
    assert closes[0]["fields"]["valid_to"] == "2026-07-02"
    created = _creates(fa)
    assert len(created) == 1
    cf = created[0]["fields"]
    assert cf["rule_key"] == _key("2026-07-02")
    assert cf["source"] == "editor:me@x", f"provenance stamp wrong: {cf.get('source')!r}"
    assert res.rule_keys_touched == [_key("2026-07-02")]


def test_price_change_does_not_close_future_dated_sibling():
    """THE close guard, re-asserted through the editor path."""
    existing = [_pr_record("rec_future", _key("2026-12-01"), valid_from="2026-12-01")]
    with FakeAirtable(existing) as fa:
        res = apply_master_change(_price_change(date(2026, 7, 2)), "me@x")
    assert res.closed == 0, f"future-dated sibling was closed (closed={res.closed})"
    assert not any(
        u["id"] == "rec_future" for u in _updates(fa)
    ), "future-dated open rule must NOT be touched"
    assert res.created == 1


def test_price_change_same_key_patches_in_place():
    """Same-key belt: re-applying the same effective date updates, not layers."""
    existing = [_pr_record("rec_same", _key("2026-07-02"), valid_from="2026-07-02")]
    with FakeAirtable(existing) as fa:
        res = apply_master_change(_price_change(date(2026, 7, 2)), "me@x")
    assert (res.created, res.updated, res.closed) == (0, 1, 0), (
        f"expected in-place PATCH (0,1,0), got {(res.created, res.updated, res.closed)}"
    )
    assert not any(set(u["fields"]) == {"valid_to"} for u in _updates(fa)), (
        "same-key rule must never be closed by its own re-apply"
    )


# --------------------------------------------------------------------------
# apply_master_change: fix_in_place / add_rule
# --------------------------------------------------------------------------

def test_fix_in_place_updates_not_creates_and_prepends_reason():
    rec = _pr_record("rec_fix", _key("2026-01-01"), valid_from="2026-01-01")
    rec["fields"]["tenant_price"] = 180.0
    rec["fields"]["reason"] = "june upload"
    change = MasterChange(
        op="fix_in_place", site_id=SITE_ID, product_code=PROD_CODE,
        tenant_price=182.0, valid_from=date(2026, 1, 1), reason="typo in upload",
    )
    with FakeAirtable([rec]) as fa:
        res = apply_master_change(change, "me@x")
    assert (res.created, res.updated, res.closed) == (0, 1, 0), (
        f"fix_in_place must PATCH in place, got {(res.created, res.updated, res.closed)}"
    )
    patched = [u for u in _updates(fa) if u["id"] == "rec_fix"]
    assert len(patched) == 1
    reason = patched[0]["fields"]["reason"]
    assert reason.startswith("corrected by me@x"), f"missing correction stamp: {reason!r}"
    assert "was £180.00" in reason, f"old figure not recorded: {reason!r}"
    assert "typo in upload" in reason
    assert reason.endswith("june upload"), (
        f"old reason must be PREPENDED to, not replaced: {reason!r}"
    )
    assert "valid_to" not in patched[0]["fields"], (
        "fix_in_place must not touch valid_to (None-strip keeps the stored value)"
    )


def test_add_rule_new_key_creates_and_autocreates_site_product():
    change = MasterChange(
        op="add_rule", site_id="999", product_code="PNEW", product_desc="New Keg",
        tenant_price=95.0, valid_from=date(2026, 7, 1), reason="new line",
        create_missing_site=True, create_missing_product=True,
    )
    with FakeAirtable([]) as fa:
        res = apply_master_change(change, "me@x")
    assert (res.created, res.updated, res.closed) == (1, 0, 0)
    site_creates = _creates(fa, airtable_io.T["Sites"])
    assert len(site_creates) == 1 and site_creates[0]["fields"] == {
        "site_id": "999", "status": "tenanted", "country": "england",
    }, f"site auto-create defaults wrong: {site_creates}"
    prod_creates = _creates(fa, airtable_io.T["Products"])
    assert len(prod_creates) == 1 and prod_creates[0]["fields"]["supplier"] == "LWC", (
        f"product auto-create defaults wrong: {prod_creates}"
    )
    created = _creates(fa)
    assert created[0]["fields"]["rule_key"] == "999|PNEW|2026-07-01"


def test_apply_end_rule_dispatches_to_end_pricing_rule():
    rec = _pr_record("rec_open", _key("2026-01-01"), valid_from="2026-01-01")
    change = MasterChange(
        op="end_rule", site_id=SITE_ID, product_code=PROD_CODE,
        valid_from=date(2026, 1, 1), valid_to=date(2026, 7, 1), reason="delist",
    )
    with FakeAirtable([rec]) as fa:
        res = apply_master_change(change, "me@x")
    assert (res.created, res.updated, res.closed) == (0, 1, 1)
    f = _updates(fa)[0]["fields"]
    assert f["valid_to"] == "2026-07-01" and f["source"] == "editor:me@x"
    assert res.rule_keys_touched == [_key("2026-01-01")]


# --------------------------------------------------------------------------
# validate_master_change — table-driven over the §2.3 invariants
# --------------------------------------------------------------------------

def _mc(**kw) -> MasterChange:
    base = dict(
        op="price_change", site_id=SITE_ID, product_code=PROD_CODE,
        tenant_price=182.0, valid_from=date(2026, 7, 2), reason="ok",
    )
    base.update(kw)
    return MasterChange(**base)


def test_validate_table():
    snap_open = _snap([_rule(vf=date(2026, 1, 1))])
    snap_ended = _snap([_rule(vf=date(2026, 1, 1), vt=date(2026, 6, 1))])
    snap_with_successor = _snap([
        _rule(vf=date(2026, 1, 1)),
        _rule(vf=date(2026, 7, 1)),  # covers the end date below
    ])
    snap_gap_successor = _snap([
        _rule(vf=date(2026, 1, 1)),
        _rule(vf=date(2026, 9, 1)),
    ])
    end_ok = dict(
        op="end_rule", tenant_price=None,
        valid_from=date(2026, 1, 1), valid_to=date(2026, 7, 1),
    )

    # (name, snap, change, expected error substring or None,
    #  expected warning substring or None)
    cases = [
        ("ok price change", snap_open, _mc(), None, None),
        ("reason required (inv 6)", snap_open, _mc(reason="  "), "reason is required", None),
        ("tenant_price required (inv 2)", snap_open, _mc(tenant_price=None), "tenant_price is required", None),
        ("tenant_price positive (inv 2)", snap_open, _mc(tenant_price=-5.0), "must be positive", None),
        ("fb_price positive (inv 2)", snap_open, _mc(fb_price=0.0), "must be positive", None),
        ("sanity band warns not blocks (inv 2)", snap_open, _mc(tenant_price=600.0), None, "sanity band"),
        ("unknown site blocks (inv 1)", snap_open, _mc(site_id="998"), "not in the master", None),
        ("unknown product blocks (inv 1)", snap_open, _mc(product_code="NOPE"), "not in the master", None),
        ("add_rule needs create opt-in (inv 1)", snap_open,
         _mc(op="add_rule", product_code="PNEW"), "tick 'create new product'", None),
        ("add_rule opt-in needs desc (inv 1)", snap_open,
         _mc(op="add_rule", product_code="PNEW", create_missing_product=True),
         "product_desc is required", None),
        ("add_rule opt-in ok, warns defaults (inv 1)", snap_open,
         _mc(op="add_rule", product_code="PNEW", product_desc="New Keg",
             create_missing_product=True), None, "auto-created"),
        ("effective date required (inv 3)", snap_open, _mc(valid_from=None),
         "effective date", None),
        ("inverted interval blocks (inv 3)", snap_open,
         _mc(valid_to=date(2026, 7, 2)), "must be after", None),
        ("key collision blocks price_change (inv 4)", snap_open,
         _mc(valid_from=date(2026, 1, 1)), "already starts on that date", None),
        ("key collision blocks add_rule (inv 4)", snap_open,
         _mc(op="add_rule", valid_from=date(2026, 1, 1)), "already starts on that date", None),
        ("collision allowed for fix_in_place (inv 4)", snap_open,
         _mc(op="fix_in_place", valid_from=date(2026, 1, 1)), None, None),
        ("fix_in_place unknown key blocks", snap_open,
         _mc(op="fix_in_place", valid_from=date(2026, 2, 2)), "no pricing rule", None),
        ("fix_in_place nothing to change", snap_open,
         _mc(op="fix_in_place", valid_from=date(2026, 1, 1), tenant_price=None),
         "nothing to change", None),
        ("fix_in_place mismatches caveat surfaced (§2.2)", snap_open,
         _mc(op="fix_in_place", valid_from=date(2026, 1, 1)), None, "duplicate mismatch"),
        ("status typo blocks pre-typecast (inv 9)", snap_open,
         _mc(status="Tenanted"), "status", None),
        ("managed warns (§2.1)", snap_open, _mc(status="managed"), None, "managed"),
        ("end_rule ok + delist warn (inv 5)", snap_open, _mc(**end_ok), None, "delist"),
        ("end_rule already ended blocks (inv 5)", snap_ended, _mc(**end_ok), "already ended", None),
        ("end_rule inverted blocks (inv 5)", snap_open,
         _mc(**{**end_ok, "valid_to": date(2026, 1, 1)}), "must be after", None),
        ("end_rule end date required (inv 5)", snap_open,
         _mc(**{**end_ok, "valid_to": None}), "end date", None),
        ("end_rule missing key blocks (inv 5)", snap_open,
         _mc(**{**end_ok, "valid_from": date(2025, 1, 1)}), "no pricing rule", None),
        ("end_rule covered by successor: no warn (inv 5)", snap_with_successor,
         _mc(**end_ok), None, None),
        ("end_rule gap to successor warns (inv 8)", snap_gap_successor,
         _mc(**end_ok), None, "gap"),
        ("add_rule after ended rule warns gap (inv 8)", snap_ended,
         _mc(op="add_rule", valid_from=date(2026, 8, 1)), None, "gap"),
        ("bad op rejected", snap_open, _mc(op="delete_rule"), "unknown op", None),
    ]

    failures = []
    for name, snap, change, err_sub, warn_sub in cases:
        errors, warnings = validate_master_change(change, snap)
        if err_sub is None:
            if errors:
                failures.append(f"{name}: unexpected errors {errors}")
        elif not any(err_sub in e for e in errors):
            failures.append(f"{name}: expected error containing {err_sub!r}, got {errors}")
        if warn_sub is not None and not any(warn_sub in w for w in warnings):
            failures.append(f"{name}: expected warning containing {warn_sub!r}, got {warnings}")
    assert not failures, "validate_master_change table failures:\n  " + "\n  ".join(failures)


def test_validate_never_forbids_overlap():
    """Overlaps are LOAD-BEARING (supports): a support layered over the open
    standard rule must validate cleanly for a different valid_from."""
    today = date.today()
    snap = _snap([
        _rule(vf=today - timedelta(days=180)),                                   # open standard
        _rule(vf=today - timedelta(days=1), vt=today + timedelta(days=30),
              status="supported"),                                                # bounded support
    ])
    change = _mc(valid_from=today + timedelta(days=7))
    errors, _warnings = validate_master_change(change, snap)
    assert not errors, f"overlapping windows must never be a blocking error: {errors}"


# --------------------------------------------------------------------------
# preview_master_change — pure
# --------------------------------------------------------------------------

def test_preview_price_change_close_create_winner():
    snap = _snap([_rule(vf=date(2026, 1, 1), tenant=180.0)])
    p = preview_master_change(_price_change(date(2026, 7, 2)), snap)
    assert not p.errors, p.errors
    assert len(p.will_close) == 1 and p.will_close[0]["valid_to"] == "2026-07-02"
    assert p.will_close[0]["rule_key"] == _key("2026-01-01")
    assert len(p.will_create) == 1 and p.will_create[0]["rule_key"] == _key("2026-07-02")
    assert "new rule wins" in p.winner_note, p.winner_note
    assert "closed at 2026-07-02" in p.summary, p.summary


def test_preview_close_pass_skips_bounded_support_and_future_rule():
    snap = _snap([
        _rule(vf=date(2026, 1, 1)),                                          # closes
        _rule(vf=date(2026, 6, 1), vt=date(2026, 8, 1), status="supported"),  # bounded: stays
        _rule(vf=date(2026, 12, 1)),                                         # future open: stays
    ])
    p = preview_master_change(_price_change(date(2026, 7, 2)), snap)
    assert [c["rule_key"] for c in p.will_close] == [_key("2026-01-01")], (
        f"only the prior open rule closes (guard mirrored), got {p.will_close}"
    )
    # On 2026-07-02 the new rule (valid_from 2026-07-02) is newest and wins.
    assert "new rule wins" in p.winner_note
    # ...but the future open rule (2026-12-01) reclaims the win from its start —
    # preview must say so, and it's a warning, not a block.
    assert not p.errors, p.errors
    assert "2026-12-01" in p.winner_note, p.winner_note
    assert any("2026-12-01" in w and "2026-07-02..2026-12-01" in w for w in p.warnings), p.warnings


def test_backdated_price_change_behind_standing_open_rule_warns_with_window():
    # The QA-found trap: change effective 2026-06-01 behind an open rule that
    # started 2026-06-15. The close pass can't close the newer rule, so the old
    # price keeps winning from 2026-06-15 — the change only bites 06-01..06-15.
    snap = _snap([_rule(vf=date(2026, 6, 15), tenant=200.0)])
    p = preview_master_change(_price_change(date(2026, 6, 1)), snap)
    assert not p.errors, p.errors  # legitimate for a scheduled future rule; warn, don't block
    assert not p.will_close, "the newer open rule must NOT be closed"
    assert any("2026-06-15" in w and "2026-06-01..2026-06-15" in w for w in p.warnings), p.warnings
    assert "2026-06-15" in p.winner_note and "only applies" in p.winner_note, p.winner_note


def test_compute_margin_math():
    from master_changes import compute_margin
    # no retro: net cost = fb; margin = tenant - fb; % of tenant
    m = compute_margin(200.0, 120.0, 0.0)
    assert m.net_cost == 120.0 and m.gross_gbp == 80.0 and m.net_gbp == 80.0
    assert abs(m.pct - 40.0) < 1e-9
    # with retro: net cost = fb*(1-retro); margin includes the rebate
    m = compute_margin(200.0, 120.0, 0.125)  # net cost 105 -> net margin 95
    assert abs(m.net_cost - 105.0) < 1e-9 and m.gross_gbp == 80.0
    assert abs(m.net_gbp - 95.0) < 1e-9 and abs(m.pct - 47.5) < 1e-9
    # selling under cost -> negative margin
    m = compute_margin(100.0, 120.0, 0.0)
    assert m.net_gbp == -20.0 and abs(m.pct - (-20.0)) < 1e-9
    # partial data / div-by-zero guards -> None, never raises
    assert compute_margin(200.0, None, 0.0).net_gbp is None
    assert compute_margin(None, 120.0, 0.0).net_gbp is None
    assert compute_margin(0.0, 120.0, 0.0).pct is None  # no divide-by-zero
    # managed-style zero margin
    assert compute_margin(120.0, 120.0, 0.0).net_gbp == 0.0


def test_retro_pct_ge_one_is_blocked():
    # retro >= the full FB list price (fraction >= 1) would make the net price
    # zero/negative — block it rather than warn-through. (The £ form converts to
    # a fraction before this runs, so this guards a mistyped/huge figure.)
    snap = _snap([_rule(vf=date(2026, 1, 1))])
    change = MasterChange(
        op="price_change", site_id=SITE_ID, product_code=PROD_CODE,
        tenant_price=100.0, fb_price=120.0, retro_pct=12.5,
        valid_from=date(2026, 7, 2), reason="typo test",
    )
    errors, _ = validate_master_change(change, snap)
    assert any(
        "at or above the FB list price" in e or "net price would be zero" in e
        for e in errors
    ), errors
    # a proper fraction is fine
    change2 = replace(change, retro_pct=0.125)
    errors2, _ = validate_master_change(change2, snap)
    assert not any("retro" in e.lower() for e in errors2), errors2


def test_render_master_pivot_shape_and_winner():
    from master_pages import render_master_pivot
    S1, S2, P1, P2 = "001", "002", "AAA", "BBB"
    def R(site, code, tenant, fb=120.0, retro=0.0, vf=None, vt=None, desc="Keg"):
        return Rule(site_id=site, product_code=code, product_desc=desc,
                    tenant_price=tenant, fb_price=fb, retro_pct=retro,
                    valid_from=vf, valid_to=vt, status="tenanted",
                    reason="x", source="test")
    rules = [
        # P1: consistent fb across both sites -> left columns show Price/Net.
        R(S1, P1, 200.0, fb=120.0, retro=0.125, vf=date(2026, 1, 1)),  # margin 95
        R(S1, P1, 999.0, fb=120.0, retro=0.125, vf=date.today() + timedelta(days=1)),  # future
        R(S2, P1, 180.0, fb=120.0, retro=0.125, vf=date(2026, 1, 1)),  # margin 75
        # P2: fb DIFFERS across sites -> Price/Net must say "varies", not pick one.
        R(S1, P2, 100.0, fb=120.0, retro=0.0, vf=date(2026, 1, 1), desc="Loss Keg"),  # margin -20
        R(S2, P2, 150.0, fb=100.0, retro=0.0, vf=date(2026, 1, 1), desc="Loss Keg"),  # margin  50
    ]
    snap = MasterSnapshot(
        sites={S1: {"name": "Alpha Arms"}, S2: {"name": "Beta Bar"}},
        rules=rules, site_ids={S1: "r1", S2: "r2"},
        product_ids={P1: "p1", P2: "p2"}, rule_ids={}, banner_info={},
        # Product-level master data (the Excel's Retro P/Keg is a fixed £):
        # P1 retro £15/keg; RONLY has a retro but no current rules -> still a row.
        products={
            P1: {"desc": "Keg", "retro_per_keg": 15.0},
            P2: {"desc": "Loss Keg", "retro_per_keg": 0.0},
            "RONLY": {"desc": "Zed Retro-Only Keg", "retro_per_keg": 7.5},
        },
    )
    html = render_master_pivot(snap, {}, is_admin=True)
    assert 'class="pivot"' in html
    # both sites are columns, both products are rows
    assert "Alpha Arms" in html and "Beta Bar" in html
    assert P1 in html and P2 in html and "Loss Keg" in html
    # Excel semantics: retro-only product still gets a row (export parity)
    assert "RONLY" in html and "Zed Retro-Only Keg" in html
    # today's winner is £200, NOT the future £999
    assert "£200.00" in html and "£999.00" not in html
    # £-band colours (operator-set): P1×S1 margin £95 -> green; P2×S2 margin £50
    # -> amber; P2×S1 margin -£20 -> red
    assert "cell-pos" in html and "cell-warn" in html and "cell-neg" in html
    # P1 left columns: Retro P/Keg = product-level £15.00; Net = 120 − 15 = £105.00
    assert "£15.00" in html and "£105.00" in html
    # P2 has different FB across sites: Price/Net must flag it, not pick one
    assert "varies" in html
    # Excel column headers + row order by product NAME (Keg < Loss Keg < Zed…)
    assert "Product Code" in html and "Retro P/Keg" in html and "Net price" in html
    assert html.index(">Keg<") < html.index("Loss Keg") < html.index("Zed Retro-Only Keg")
    # read-only by default: no edit affordances without ?edit=1
    assert 'class="cellf"' not in html and "cell-input" not in html
    # back-to-LWC link present; upload-provenance banner is NOT rendered here
    assert "Back to LWC" in html


def test_render_master_pivot_cask_section_and_edit_mode():
    from master_pages import render_master_pivot
    S1, DR, CK = "001", "ZZZ", "AAA"
    def R(site, code, tenant, desc, fb=120.0):
        return Rule(site_id=site, product_code=code, product_desc=desc,
                    tenant_price=tenant, fb_price=fb, retro_pct=0.0,
                    valid_from=date(2026, 1, 1), valid_to=None, status="tenanted",
                    reason="x", source="test")
    # The cask product's name sorts FIRST alphabetically ("Aardvark…") but must
    # still land AFTER every draught product (Excel section order).
    S2 = "002"
    rules = [
        R(S1, DR, 200.0, "Zebra Lager 50L"),
        R(S1, CK, 195.0, "Aardvark Cask 9G"),
        R(S2, DR, 190.0, "Zebra Lager 50L"),  # CK×S2 has no price -> '+' in edit mode
    ]
    snap = MasterSnapshot(
        sites={S1: {"name": "Alpha Arms"}, S2: {"name": "Beta Bar"}},
        rules=rules, site_ids={S1: "r1", S2: "r2"},
        product_ids={DR: "p1", CK: "p2"}, rule_ids={}, banner_info={},
        products={DR: {"desc": "Zebra Lager 50L", "retro_per_keg": 0.0},
                  CK: {"desc": "Aardvark Cask 9G", "retro_per_keg": 0.0}},
    )
    html = render_master_pivot(snap, {}, is_admin=True)
    assert html.index("Zebra Lager") < html.index("Aardvark Cask"), (
        "draught must come before cask regardless of alphabetical order"
    )
    # section divider rows present when a cask section exists
    assert ">Draught<" in html and ">Cask<" in html
    # edit mode (admin): priced cell links to the edit form, and a second site
    # would show '+' add links; the add-product button appears
    html_e = render_master_pivot(snap, {"edit": "1"}, is_admin=True)
    # in-grid editing: every cell is a one-field form posting to /master/cell/apply
    assert 'class="cellf"' in html_e and "/master/cell/apply" in html_e
    assert 'name="tenant_price"' in html_e and "cell-input" in html_e
    assert "Add product" in html_e and "Done editing" in html_e
    # priced DR×S1 cell prefilled + prev-tracked (clearing it = remove)
    assert 'value="200.00" data-prev="200.00"' in html_e
    # blank CK×S2 cell -> EMPTY input (type to add), identity in hidden fields
    assert 'value="" data-prev=""' in html_e
    assert 'name="site_id" value="002"' in html_e
    # site selector present, and filtering to one site drops the other column
    assert '<select name="site"' in html_e
    html_s1 = render_master_pivot(snap, {"site": S1}, is_admin=True)
    thead_s1 = html_s1[html_s1.index("<thead>"):html_s1.index("</thead>")]
    assert "Alpha Arms" in thead_s1 and "Beta Bar" not in thead_s1, (
        "site filter must show ONLY the chosen site's column"
    )
    # single-site view stays in the page column (no full-bleed breakout)
    assert 'class="pivot-single"' in html_s1 and 'class="pivot-wide"' not in html_s1
    html_all = render_master_pivot(snap, {}, is_admin=True)
    assert 'class="pivot-wide"' in html_all
    # edit mode is admin-only: a viewer passing ?edit=1 gets the read-only grid
    html_v = render_master_pivot(snap, {"edit": "1"}, is_admin=False)
    assert 'class="cellf"' not in html_v and "cell-input" not in html_v
    assert "Edit prices" not in html_v


def test_master_banner_escapes_source_filename():
    # The master banner shows the uploaded FILENAME (Rule.source) — an
    # attacker-influenceable value. It must be HTML-escaped, not injected.
    import webapp
    html = webapp._render_master_banner({
        "sources": ["<script>alert(1)</script>evil.xlsx"],
        "latest_valid_from": "<img src=x onerror=alert(1)>",
        "latest_uploaded_at": "2026-07-02T10:00",
        "active_rule_count": 3, "products_with_retro": 1,
    })
    assert "<script>" not in html and "&lt;script&gt;" in html
    assert "<img src=x" not in html


def test_retro_gbp_form_converts_to_fraction():
    # The edit/add forms collect retro as £/keg (retro_gbp); the codec stores it
    # as a fraction of the FB list price. £22.50 on a £180 list -> 0.125.
    from master_pages import parse_master_change_form
    base = {
        "op": "price_change", "site_id": SITE_ID, "product_code": PROD_CODE,
        "tenant_price": "200", "fb_price": "180", "retro_gbp": "22.50",
        "status": "tenanted", "valid_from": "2026-07-02", "reason": "retro £ test",
    }
    change, errs = parse_master_change_form(base)
    assert not errs, errs
    assert change.retro_pct is not None and abs(change.retro_pct - 0.125) < 1e-12

    # The confirm→apply path carries the exact fraction (retro_pct hidden field);
    # it must win over any stale £ field so the round-trip stays bit-exact.
    both = {**base, "retro_pct": "0.1"}
    change2, _ = parse_master_change_form(both)
    assert change2.retro_pct == 0.1

    # £ retro with no list price to divide by -> a loud error, never a silent drop.
    no_fb = {k: v for k, v in base.items() if k != "fb_price"}
    change3, errs3 = parse_master_change_form(no_fb)
    assert any("FB list price" in e for e in errs3), errs3

    # Blank retro -> None (the "keep existing" / "no retro" path), no error.
    blank = {**base, "retro_gbp": ""}
    change4, errs4 = parse_master_change_form(blank)
    assert change4.retro_pct is None and not errs4


def test_preview_winner_overlapping_support_wins_today():
    """§5.2: winner-on-date with an overlapping support rule — newest
    valid_from first, mirroring reconcile._index_rules."""
    today = date.today()
    std_vf = today - timedelta(days=180)
    sup_vf = today - timedelta(days=1)
    snap = _snap([
        _rule(vf=std_vf, tenant=180.0),
        _rule(vf=sup_vf, vt=today + timedelta(days=30), tenant=150.0, status="supported"),
    ])
    change = MasterChange(
        op="fix_in_place", site_id=SITE_ID, product_code=PROD_CODE,
        tenant_price=182.0, valid_from=std_vf, reason="fix",
    )
    p = preview_master_change(change, snap)
    sup_key = _key(sup_vf.isoformat())
    assert sup_key in p.winner_note and "wins" in p.winner_note, (
        f"the overlapping support (newest valid_from) must be named the winner: {p.winner_note!r}"
    )


def test_preview_end_rule_delist_and_takeover():
    # No successor: delist warning + membership note.
    snap = _snap([_rule(vf=date(2026, 1, 1))])
    change = MasterChange(
        op="end_rule", site_id=SITE_ID, product_code=PROD_CODE,
        valid_from=date(2026, 1, 1), valid_to=date(2026, 7, 1), reason="delist",
    )
    p = preview_master_change(change, snap)
    assert not p.errors, p.errors
    assert len(p.will_update) == 1 and p.will_update[0]["new"]["valid_to"] == "2026-07-01"
    assert not p.will_create and not p.will_close, "end_rule creates/closes nothing else"
    assert "drops off" in p.winner_note, p.winner_note
    assert any("delist" in w for w in p.warnings), p.warnings
    # With a covering successor: it takes over, no delist warning.
    snap2 = _snap([_rule(vf=date(2026, 1, 1)), _rule(vf=date(2026, 7, 1))])
    p2 = preview_master_change(change, snap2)
    assert _key("2026-07-01") in p2.winner_note and "takes over" in p2.winner_note
    assert not any("delist" in w for w in p2.warnings), p2.warnings


def test_change_rule_key_normalises_codes():
    # _to_str_code: floats lose the trailing .0 (Excel artefact), strings are stripped.
    change = _mc(site_id=809.0, product_code=" PKEG1 ", valid_from=None)
    assert change_rule_key(change) == "809|PKEG1|open"


# --------------------------------------------------------------------------
# Route smoke tests (design §5.2 last bullet) — FastAPI TestClient, offline.
#
# The TestClient is deliberately NOT used as a context manager: starlette only
# runs lifespan (and webapp's master-cache warm-up thread, which would fetch
# through whatever _list_all is installed at the time) inside `with`; skipping
# it keeps requests strictly inside the FakeAirtable window.
# --------------------------------------------------------------------------

def _reset_master_cache() -> None:
    """Blow away the real module-level snapshot cache so each route test reads
    through the CURRENT fakes (FakeAirtable's patched invalidate_master_cache
    is a counter and does not clear the real cache)."""
    with airtable_io._MASTER_CACHE_LOCK:
        airtable_io._MASTER_CACHE["snapshot"] = None
        airtable_io._MASTER_CACHE["ts"] = 0.0
        # Also clear the stale-fallback snapshot, else one test's master leaks
        # into the next (production retains "last" across writes on purpose; a
        # test swapping fake fixtures must not).
        airtable_io._MASTER_CACHE["last"] = None
        airtable_io._MASTER_CACHE["gen"] += 1


class FakeAuthClient:
    """Context manager yielding a TestClient with:
      - auth stubbed: auth_supabase._resolve_supabase returns a principal of
        the given role (the require_drinks_role closure looks it up at call
        time in the auth_supabase module namespace);
      - webapp.refresh_master_cache_async disarmed (a background rebuild
        thread could outlive the FakeAirtable window and hit the network);
      - the real master cache reset on entry and exit.
    """

    EMAIL = "tester@example.com"

    def __init__(self, role: str):
        self.role = role

    def __enter__(self):
        import auth_supabase
        import webapp
        from fastapi.testclient import TestClient

        self._auth = auth_supabase
        self._webapp = webapp
        self._orig_resolve = auth_supabase._resolve_supabase
        self._orig_refresh = webapp.refresh_master_cache_async
        role, email = self.role, self.EMAIL
        auth_supabase._resolve_supabase = lambda request: (
            auth_supabase.DrinksPrincipal(email=email, role=role),
            None,
        )
        webapp.refresh_master_cache_async = lambda: None
        _reset_master_cache()
        # raise_server_exceptions=False: unhandled errors run through webapp's
        # own 500 handler (as in prod) instead of failing the test harness.
        return TestClient(webapp.app, raise_server_exceptions=False)

    def __exit__(self, *exc):
        self._auth._resolve_supabase = self._orig_resolve
        self._webapp.refresh_master_cache_async = self._orig_refresh
        _reset_master_cache()
        return False


def _extract_hidden(html_text: str) -> dict[str, str]:
    """Pull the confirm page's hidden inputs (we control the renderer, so a
    regex over the fixed attribute order is reliable)."""
    return {
        unescape(m.group(1)): unescape(m.group(2))
        for m in re.finditer(
            r'<input type="hidden" name="([^"]+)" value="([^"]*)"', html_text
        )
    }


def _grid_rule(rec_id: str = "rec_open", vf: str = "2026-01-01") -> dict:
    rec = _pr_record(rec_id, _key(vf), valid_from=vf)
    rec["fields"]["tenant_price"] = 180.0
    rec["fields"]["status"] = "tenanted"
    return rec


def test_mismatch_writer_is_idempotent_on_reupload():
    """Re-uploading the same supplier file (deduped by content hash -> same
    deterministic mismatch_keys) must NOT duplicate mismatch rows — the fix for
    a bookkeeper re-uploading after a proxy timeout."""
    def payload():
        return [
            {"fields": {"mismatch_key": "fileA|0001|001|PKEG1|INV1|wrong_price", "type": "x", "status": "open"}},
            {"fields": {"mismatch_key": "fileA|0002|001|PKEG2|INV2|wrong_price", "type": "y", "status": "open"}},
        ]
    with FakeAirtable([]) as fa:
        fa.tables[airtable_io.T["Mismatches"]] = []  # empty findings table
        n1 = airtable_io._create_mismatches_deduped(payload())
        n2 = airtable_io._create_mismatches_deduped(payload())  # re-upload
    assert n1 == 2, "first upload writes both findings"
    assert n2 == 0, "re-upload of the same file writes nothing (no duplicates)"
    creates = [
        r for op, t, recs in fa.calls
        if op == "create" and t == airtable_io.T["Mismatches"] for r in recs
    ]
    assert len(creates) == 2, "only the first upload's two rows were ever created"


def test_set_product_retro_updates_product_and_reflows_rules():
    """Retro is product-level (operator decision): set_product_retro PATCHes
    Products.retro_per_keg AND reflows every current tenanted rule with
    retro_pct = retro£ / that rule's fb, tenant prices unchanged."""
    from datetime import date as _date
    rec = _grid_rule("rec_old", vf="2026-01-01")
    rec["fields"]["fb_price"] = 120.0   # retro £15 -> retro_pct 0.125
    with FakeAirtable([rec]) as fa:
        n = airtable_io.set_product_retro(PROD_CODE, 15.0, _date(2026, 7, 4), "test")
    assert n == 1
    # Products PATCH carries the £ figure + eligibility
    prod_ups = [
        recs for op, t, recs in fa.calls
        if op == "update" and t == airtable_io.T["Products"]
    ]
    assert prod_ups and prod_ups[0][0]["fields"]["retro_per_keg"] == 15.0
    assert prod_ups[0][0]["fields"]["retro_eligible"] is True
    # a successor rule reflowed from today with retro_pct = 15/120 = 0.125,
    # tenant price unchanged
    created = [
        r for op, t, recs in fa.calls if op == "create" and t == airtable_io.T["PricingRules"]
        for r in recs
    ]
    assert len(created) == 1
    f = created[0]["fields"]
    assert f["tenant_price"] == 180.0
    assert abs(f["retro_pct"] - 0.125) < 1e-9
    assert f["valid_from"] == "2026-07-04"


def test_route_grid_viewer_ok_edit_admin_only():
    """Viewer can GET /master (no edit links); mutating pages are admin-gated;
    admin sees the links and the edit page renders both forms."""
    with FakeAirtable([_grid_rule()]):
        with FakeAuthClient("viewer") as client:
            # Default /master = the Excel-style pivot (read-only, no edit links).
            r = client.get("/master")
            assert r.status_code == 200, r.text[:300]
            assert PROD_CODE in r.text
            assert 'class="pivot"' in r.text, "default /master must be the pivot view"
            # upload-provenance banner removed from the pivot (list view keeps it)
            assert "Current master:" not in r.text
            assert "Back to LWC" in r.text
            # The detailed list view hides edit links from viewers too.
            rl = client.get("/master", params={"view": "list"})
            assert rl.status_code == 200 and PROD_CODE in rl.text
            assert "/master/edit" not in rl.text, "edit links must be hidden from viewers"
            r2 = client.get("/master/edit", params={"rule_key": _key("2026-01-01")})
            assert r2.status_code == 403, f"viewer must get 403 on /master/edit, got {r2.status_code}"
        with FakeAuthClient("admin") as client:
            # Pivot is read-only for everyone in Phase 1; edit links live on the list.
            rp = client.get("/master")
            assert rp.status_code == 200 and 'class="pivot"' in rp.text
            r3 = client.get("/master", params={"view": "list"})
            assert r3.status_code == 200 and "/master/edit" in r3.text
            r4 = client.get("/master/edit", params={"rule_key": _key("2026-01-01")})
            assert r4.status_code == 200
            assert "Change price from a date" in r4.text and "Fix a mistake" in r4.text
            # status prefill: a price-only fix must not silently reset status
            assert '<option value="tenanted" selected>' in r4.text


def test_route_end_add_forms_and_404():
    with FakeAirtable([_grid_rule()]):
        with FakeAuthClient("admin") as client:
            r = client.get("/master/end", params={"rule_key": _key("2026-01-01")})
            assert r.status_code == 200 and "delist" in r.text.lower()
            r2 = client.get("/master/add")
            assert r2.status_code == 200 and "create this product" in r2.text.lower()
            r3 = client.get("/master/edit", params={"rule_key": "999|NOPE|open"})
            assert r3.status_code == 404, f"stale rule_key must 404 politely, got {r3.status_code}"


def test_route_preview_apply_roundtrip():
    """§5.2: admin preview→apply carries the change intact through the hidden
    inputs; preview writes NOTHING; apply closes the prior rule, creates the
    successor, and stamps provenance from the signed-in principal."""
    form = {
        "op": "price_change", "site_id": SITE_ID, "product_code": PROD_CODE,
        "tenant_price": "182.0", "status": "tenanted",
        "valid_from": "2026-07-02", "reason": "LWC list increase Jul-26",
    }
    with FakeAirtable([_grid_rule("rec_old")]) as fa:
        with FakeAuthClient("admin") as client:
            r = client.post("/master/preview", data=form)
            assert r.status_code == 200, r.text[:500]
            assert "Confirm change" in r.text
            assert not fa.calls, f"preview must write NOTHING, got {fa.calls}"
            hidden = _extract_hidden(r.text)
            assert hidden["op"] == "price_change"
            assert hidden["tenant_price"] == "182.0"
            assert hidden["valid_from"] == "2026-07-02"
            assert hidden["reason"] == "LWC list increase Jul-26", (
                f"change must round-trip the confirm page intact: {hidden}"
            )
            r2 = client.post("/master/apply", data=hidden)
            assert r2.status_code == 200, r2.text[:500]
    closes = [u for u in _updates(fa) if set(u["fields"]) == {"valid_to"}]
    assert [u["id"] for u in closes] == ["rec_old"]
    assert closes[0]["fields"]["valid_to"] == "2026-07-02"
    created = _creates(fa)
    assert len(created) == 1 and created[0]["fields"]["rule_key"] == _key("2026-07-02")
    assert created[0]["fields"]["source"] == f"editor:{FakeAuthClient.EMAIL}", (
        f"provenance stamp wrong: {created[0]['fields'].get('source')!r}"
    )
    assert "Rules created" in r2.text and "Back to price grid" in r2.text


def test_route_cell_editor_amend_and_remove():
    """The Excel-like cell editor: GET renders the tiny form; POST save closes
    the old rule + creates today's successor INHERITING fb/retro/status
    server-side (client sends only the price); the redirect returns to the
    grid with the PATCHED cache already showing the new price; POST delete
    ends the rule. Viewer role is locked out."""
    today_iso = date.today().isoformat()
    rec = _grid_rule("rec_old")
    rec["fields"]["fb_price"] = 200.0
    with FakeAirtable([rec]) as fa:
        with FakeAuthClient("admin") as client:
            r = client.get("/master/cell", params={
                "site_id": SITE_ID, "product_code": PROD_CODE,
            })
            assert r.status_code == 200 and "Change price" in r.text
            assert 'name="tenant_price"' in r.text and "Remove price" in r.text

            r2 = client.post("/master/cell/apply", data={
                "site_id": SITE_ID, "product_code": PROD_CODE,
                "do": "save", "tenant_price": "199.50", "fsite": SITE_ID,
            }, follow_redirects=False)
            assert r2.status_code == 303, r2.text[:300]
            loc = r2.headers["location"]
            assert "saved=1" in loc and "edit=1" in loc and f"site={SITE_ID}" in loc

            closes = [u for u in _updates(fa) if set(u["fields"]) == {"valid_to"}]
            assert [u["id"] for u in closes] == ["rec_old"]
            assert closes[0]["fields"]["valid_to"] == today_iso
            created = _creates(fa)
            assert len(created) == 1
            f = created[0]["fields"]
            assert f["tenant_price"] == 199.5
            assert f["fb_price"] == 200.0, "fb must be inherited server-side"
            assert f["source"] == f"grid:{FakeAuthClient.EMAIL}"

            # The redirect target reads the PATCHED cache: new price visible
            # immediately (as the edit-mode cell input's value), no inline
            # Airtable sweep; and in the read-only grid as £ text.
            r3 = client.get("/master", params={"edit": "1", "saved": "1"})
            assert 'value="199.50"' in r3.text and "Saved" in r3.text
            r3v = client.get("/master")
            assert "£199.50" in r3v.text

            # Excel semantics: re-entering the SAME price is a silent no-op —
            # redirect without "saved" and NOTHING written.
            calls_before = len(fa.calls)
            r3b = client.post("/master/cell/apply", data={
                "site_id": SITE_ID, "product_code": PROD_CODE,
                "do": "save", "tenant_price": "199.50",
            }, follow_redirects=False)
            assert r3b.status_code == 303 and "saved=1" not in r3b.headers["location"]
            assert len(fa.calls) == calls_before, "same-price save must write NOTHING"

            # Excel semantics: an EMPTIED cell removes the price (winner
            # started today -> ends tomorrow, bills today only).
            r4 = client.post("/master/cell/apply", data={
                "site_id": SITE_ID, "product_code": PROD_CODE,
                "do": "save", "tenant_price": "",
            }, follow_redirects=False)
            assert r4.status_code == 303, r4.text[:300]

        with FakeAuthClient("viewer") as client:
            r5 = client.get("/master/cell", params={
                "site_id": SITE_ID, "product_code": PROD_CODE,
            })
            assert r5.status_code == 403
            r6 = client.post("/master/cell/apply", data={
                "site_id": SITE_ID, "product_code": PROD_CODE,
                "do": "save", "tenant_price": "1.00",
            }, follow_redirects=False)
            assert r6.status_code == 403


def test_route_site_rename():
    """Grid edit-mode site headers link to /master/site; POST renames the site
    in Airtable (Sites PATCH), the patched cache shows the new name on the
    redirect target immediately, and viewers are locked out."""
    with FakeAirtable([_grid_rule()]) as fa:
        with FakeAuthClient("admin") as client:
            # edit-mode header links to the rename form
            r0 = client.get("/master", params={"edit": "1"})
            assert "/master/site" in r0.text and "site-head" in r0.text
            r = client.get("/master/site", params={"site_id": SITE_ID})
            assert r.status_code == 200 and 'name="name"' in r.text

            r2 = client.post("/master/site/apply", data={
                "site_id": SITE_ID, "name": "Victoria Barnsley",
            }, follow_redirects=False)
            assert r2.status_code == 303, r2.text[:300]
            ups = [
                (op, recs) for op, t, recs in fa.calls
                if op == "update" and t == airtable_io.T["Sites"]
            ]
            assert len(ups) == 1 and ups[0][1][0]["fields"] == {"name": "Victoria Barnsley"}
            assert ups[0][1][0]["id"] == SITE_REC

            # patched cache: the new name is in the header immediately
            r3 = client.get("/master")
            assert "Victoria Barnsley" in r3.text

            # empty name -> 400, nothing written
            n_calls = len(fa.calls)
            r4 = client.post("/master/site/apply", data={
                "site_id": SITE_ID, "name": "  ",
            }, follow_redirects=False)
            assert r4.status_code == 400 and len(fa.calls) == n_calls

        with FakeAuthClient("viewer") as client:
            assert client.get("/master/site", params={"site_id": SITE_ID}).status_code == 403
            r5 = client.post("/master/site/apply", data={
                "site_id": SITE_ID, "name": "X",
            }, follow_redirects=False)
            assert r5.status_code == 403


def test_build_universal_increase_math():
    """The annual increase: tenant and FB list rise by pct; the retro stays a
    FIXED £/keg (retro_pct is recomputed against the new list); supported/
    managed, future-dated, ended and price-less rules are left alone."""
    from master_changes import build_universal_increase
    S1 = SITE_ID
    def R(code, tenant, fb, retro_pct=0.0, status="tenanted", vf=date(2026, 1, 1), vt=None):
        return Rule(site_id=S1, product_code=code, product_desc=code,
                    tenant_price=tenant, fb_price=fb, retro_pct=retro_pct,
                    valid_from=vf, valid_to=vt, status=status, reason="x", source="s")
    rules = [
        R("AAA", 200.0, 120.0, retro_pct=0.125),                  # included
        R("BBB", 100.0, None),                                    # included, no fb
        R("SUP", 90.0, 120.0, status="supported"),                # skipped: support
        R("FUT", 100.0, 120.0, vf=date(2099, 1, 1)),              # skipped: future
        R("END", 100.0, 120.0, vt=date(2026, 6, 1)),              # skipped: ended
        R("NOP", None, 120.0),                                    # skipped: no tenant price
    ]
    snap = MasterSnapshot(
        sites={S1: {"name": "T"}}, rules=rules, site_ids={S1: "r"},
        product_ids={c: c for c in ("AAA", "BBB", "SUP", "FUT", "END", "NOP")},
        rule_ids={}, banner_info={},
    )
    d = date(2026, 7, 3)
    new_rules, stats = build_universal_increase(snap, 3.5, d, "me@x")
    assert stats["n_rules"] == 2 and len(new_rules) == 2
    assert stats["skipped_support"] == 1 and stats["skipped_future"] == 1
    assert stats["skipped_no_price"] == 1
    by_code = {r.product_code: r for r in new_rules}
    a = by_code["AAA"]
    assert a.tenant_price == 207.0                      # 200 * 1.035
    assert a.fb_price == 124.2                          # 120 * 1.035
    # retro £ preserved EXACTLY: was 0.125*120 = 15.00 -> 15.00 on the new list
    assert abs(a.retro_pct * a.fb_price - 15.0) < 1e-9
    assert a.valid_from == d and a.valid_to is None and a.status == "tenanted"
    assert a.source == "increase:me@x"
    b = by_code["BBB"]
    assert b.tenant_price == 103.5 and b.fb_price is None


def test_route_universal_increase_preview_apply():
    """Preview writes NOTHING; apply closes the old rule and creates the
    successor via the bulk upsert; the patched cache shows the new price."""
    rec = _grid_rule("rec_old")           # tenant 180.0, vf 2026-01-01
    rec["fields"]["fb_price"] = 120.0
    with FakeAirtable([rec]) as fa:
        with FakeAuthClient("admin") as client:
            r = client.post("/master/increase/preview", data={
                "pct": "5", "valid_from": date.today().isoformat(),
            })
            assert r.status_code == 200 and "Confirm: +5%" in r.text
            assert "189.00" in r.text, "preview must show 180 -> 189"
            assert not fa.calls, f"preview must write NOTHING, got {fa.calls}"
            hidden = _extract_hidden(r.text)
            assert hidden.get("state"), "preview must carry the idempotence fingerprint"

            r2 = client.post("/master/increase/apply", data=hidden)
            assert r2.status_code == 200 and "Increase applied" in r2.text
            closes = [u for u in _updates(fa) if "valid_to" in u["fields"]]
            assert [u["id"] for u in closes] == ["rec_old"]
            created = _creates(fa)
            assert len(created) == 1 and created[0]["fields"]["tenant_price"] == 189.0
            assert created[0]["fields"]["fb_price"] == 126.0
            assert created[0]["fields"]["source"] == f"increase:{FakeAuthClient.EMAIL}"

            # patched cache: the grid shows the new price immediately
            r3 = client.get("/master")
            assert "£189.00" in r3.text and "£180.00" not in r3.text

            # DOUBLE SUBMIT: re-posting the same confirm must be refused with
            # nothing further written — otherwise +5% would compound to +10.25%.
            calls_before = len(fa.calls)
            r3b = client.post("/master/increase/apply", data=hidden)
            assert r3b.status_code == 400 and "already applied" in r3b.text
            assert len(fa.calls) == calls_before, "double submit must write NOTHING"

            # fat-finger guard: ±50% hard limit blocks at the form
            r4 = client.post("/master/increase/preview", data={
                "pct": "500", "valid_from": date.today().isoformat(),
            })
            assert r4.status_code == 400 and "hard limit" in r4.text
        with FakeAuthClient("viewer") as client:
            assert client.get("/master/increase").status_code == 403


def test_route_site_create_end_all_delete():
    """Add-site creates the record and lands on its empty column; end_all ends
    every open rule at a site; delete refuses when history exists and succeeds
    for a rule-less site."""
    with FakeAirtable([_grid_rule("rec_old")]) as fa:
        with FakeAuthClient("admin") as client:
            # create a brand-new site
            r = client.post("/master/site/create", data={
                "site_id": "830", "name": "Manor House",
            }, follow_redirects=False)
            assert r.status_code == 303 and "site=830" in r.headers["location"]
            creates = [
                (t, recs) for op, t, recs in fa.calls
                if op == "create" and t == airtable_io.T["Sites"]
            ]
            assert len(creates) == 1
            assert creates[0][1][0]["fields"]["site_id"] == "830"
            assert creates[0][1][0]["fields"]["name"] == "Manor House"
            # the patched cache offers the new site in edit mode immediately
            r1b = client.get("/master", params={"edit": "1"})
            assert "Manor House" in r1b.text

            # duplicate id refused
            r2 = client.post("/master/site/create", data={
                "site_id": SITE_ID, "name": "Dup",
            }, follow_redirects=False)
            assert r2.status_code == 400 and "already exists" in r2.text

            # delete refused while pricing history exists
            r3 = client.post("/master/site/apply", data={
                "site_id": SITE_ID, "do": "delete",
            }, follow_redirects=False)
            assert r3.status_code == 400 and "history" in r3.text

            # end_all closes the open rule
            r4 = client.post("/master/site/apply", data={
                "site_id": SITE_ID, "do": "end_all",
            }, follow_redirects=False)
            assert r4.status_code == 303
            ends = [
                u for u in _updates(fa)
                if "valid_to" in u["fields"] and "reason" in u["fields"]
            ]
            assert [u["id"] for u in ends] == ["rec_old"]
            assert "site removed from master" in ends[0]["fields"]["reason"]
        with FakeAuthClient("viewer") as client:
            assert client.get("/master/site/new").status_code == 403


def test_route_product_settings_rename_end_all_delete():
    """Product settings: a code change PATCHes Products AND rewrites the
    stored rule_key on every linked rule (else the next upload would duplicate
    instead of supersede); end_all delists the line; delete refuses with
    history; the patched cache shows the rename immediately."""
    with FakeAirtable([_grid_rule("rec_old")]) as fa:
        with FakeAuthClient("admin") as client:
            # edit-mode product cells link to the settings page
            r0 = client.get("/master", params={"edit": "1"})
            assert "/master/product" in r0.text
            r = client.get("/master/product", params={"product_code": PROD_CODE})
            assert r.status_code == 200 and 'name="new_code"' in r.text

            # rename code + name -> Products PATCH + rule_key rewrite
            r2 = client.post("/master/product/apply", data={
                "product_code": PROD_CODE, "do": "save",
                "new_code": "PKEG2", "new_desc": "Renamed Keg",
            }, follow_redirects=False)
            assert r2.status_code == 303, r2.text[:300]
            prod_ups = [
                recs for op, t, recs in fa.calls
                if op == "update" and t == airtable_io.T["Products"]
            ]
            assert len(prod_ups) == 1 and prod_ups[0][0]["fields"] == {
                "product_code": "PKEG2", "description": "Renamed Keg",
            }
            key_ups = [
                recs for op, t, recs in fa.calls
                if op == "update" and t == airtable_io.T["PricingRules"]
            ]
            assert len(key_ups) == 1 and key_ups[0][0]["fields"]["rule_key"] == (
                f"{SITE_ID}|PKEG2|2026-01-01"
            ), "a code change MUST rewrite the stored rule_key"

            # patched cache: grid shows the new name/code immediately
            r3 = client.get("/master")
            assert "Renamed Keg" in r3.text and "PKEG2" in r3.text

    with FakeAirtable([_grid_rule("rec_old")]) as fa_fb:
        with FakeAuthClient("admin") as client:
            # product-level FB list + retro edit: Products PATCH carries the
            # retro; every current price re-dated from today with the new
            # figures, tenant UNCHANGED (only the cost side moves).
            r = client.post("/master/product/apply", data={
                "product_code": PROD_CODE, "do": "save",
                "new_code": PROD_CODE, "new_desc": "Test Keg",
                "new_fb": "130.00", "new_retro": "13.00",
            }, follow_redirects=False)
            assert r.status_code == 303, r.text[:300]
            prod_ups = [
                recs for op, t, recs in fa_fb.calls
                if op == "update" and t == airtable_io.T["Products"]
            ]
            assert prod_ups and prod_ups[0][0]["fields"]["retro_per_keg"] == 13.0
            created = _creates(fa_fb)
            assert len(created) == 1
            f = created[0]["fields"]
            assert f["tenant_price"] == 180.0, "tenant price must NOT change"
            assert f["fb_price"] == 130.0
            assert abs(f["retro_pct"] - 13.0 / 130.0) < 1e-9
            closes = [u for u in _updates(fa_fb)
                      if u["fields"].get("valid_to") and "rule_key" not in u["fields"]]
            assert any(u["id"] == "rec_old" for u in closes), "old rule closed at today"
            # patched cache: net price 130-13=£117.00 visible immediately
            r2 = client.get("/master")
            assert "£117.00" in r2.text and "£130.00" in r2.text

    # Fresh fakes per op below — the recording fakes don't mutate their
    # tables, so each op runs against the original PKEG1 state (as Airtable
    # would hold it before that op).
    prods2 = PRODUCTS + [{"id": "rec_prod_2", "fields": {
        "product_code": "OTHER", "description": "Other Keg"}}]
    with FakeAirtable([_grid_rule("rec_old")], products=prods2):
        with FakeAuthClient("admin") as client:
            # renaming onto ANOTHER product's code is refused
            r4 = client.post("/master/product/apply", data={
                "product_code": PROD_CODE, "do": "save",
                "new_code": "OTHER", "new_desc": "Clash",
            }, follow_redirects=False)
            assert r4.status_code == 400 and "already taken" in r4.text

            # delete refused while history exists
            r5 = client.post("/master/product/apply", data={
                "product_code": PROD_CODE, "do": "delete",
            }, follow_redirects=False)
            assert r5.status_code == 400 and "history" in r5.text

    with FakeAirtable([_grid_rule("rec_old")]) as fa2:
        with FakeAuthClient("admin") as client:
            # end_all closes the open rule estate-wide
            r6 = client.post("/master/product/apply", data={
                "product_code": PROD_CODE, "do": "end_all",
            }, follow_redirects=False)
            assert r6.status_code == 303
            ends = [u for u in _updates(fa2) if "valid_to" in u["fields"]]
            assert any(u["id"] == "rec_old" for u in ends)
            assert any("product removed from master" in u["fields"].get("reason", "")
                       for u in ends)
        with FakeAuthClient("viewer") as client:
            assert client.get(
                "/master/product", params={"product_code": PROD_CODE}
            ).status_code == 403


def test_route_upload_master_updates_grid_immediately():
    """/upload-master (whole-file replace): the prior rule is closed at the
    effective date, the file's prices land, AND the very next /master read
    shows them — served from the PATCHED cache, not a post-invalidate inline
    sweep (the hub-proxy-timeout class)."""
    from io import BytesIO
    from openpyxl import Workbook

    # Cost file in the exact parse_fb_cost_file shape: row1 blank, row2
    # headers with per-site columns from col 5, data from row3.
    wb = Workbook()
    ws = wb.active
    ws.append([None] * 6)
    ws.append(["Product Code", "Product Name", "Price", "Retro P/Keg", "Net price",
               f"Test Tavern {SITE_ID}"])
    ws.append([PROD_CODE, "Test Keg", 130.0, 13.0, 117.0, 205.0])
    buf = BytesIO()
    wb.save(buf)

    with FakeAirtable([_grid_rule("rec_old")]) as fa:   # current price £180
        with FakeAuthClient("admin") as client:
            r = client.post(
                "/upload-master",
                data={"valid_from": date.today().isoformat(), "reason": "jul list"},
                files={"file": ("cost_jul.xlsx", buf.getvalue(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert r.status_code == 200, r.text[:400]
            assert "Master uploaded" in r.text and "View price grid" in r.text
            closes = [u for u in _updates(fa) if "valid_to" in u["fields"]]
            assert any(u["id"] == "rec_old" for u in closes), "old rule must be closed"
            created = _creates(fa)
            tenants = [c["fields"].get("tenant_price") for c in created
                       if c["fields"].get("tenant_price") is not None]
            assert tenants == [205.0]

            # The point: the grid ALREADY shows the file's prices (patched cache).
            r2 = client.get("/master")
            assert "£205.00" in r2.text and "£180.00" not in r2.text
            # product-level columns reflect the file's Price/Retro/Net too
            assert "£130.00" in r2.text and "£13.00" in r2.text and "£117.00" in r2.text


def test_route_export_master_from_snapshot():
    """/export-master builds from the CACHED snapshot (sub-second — the old
    fresh-sweep build timed out the hub proxy) and the workbook carries the
    current prices in the cost-file layout."""
    from io import BytesIO
    from openpyxl import load_workbook

    rec = _grid_rule("rec_open")
    rec["fields"]["fb_price"] = 120.0
    with FakeAirtable([rec]):
        with FakeAuthClient("viewer") as client:   # download is viewer-level
            r = client.get("/export-master")
            assert r.status_code == 200, r.text[:300]
            assert "spreadsheetml" in r.headers["content-type"]
            wb = load_workbook(BytesIO(r.content))
            ws = wb.active
            headers = [c.value for c in ws[2]]
            assert headers[:5] == ["Product Code", "Product Name", "Price",
                                   "Retro P/Keg", "Net price"]
            assert any(SITE_ID in str(h) for h in headers[5:]), headers
            row = [c.value for c in ws[3]]
            assert row[0] == PROD_CODE and row[2] == 120.0 and row[5] == 180.0


def test_route_apply_revalidates_tampered_hidden_inputs():
    """/master/apply must re-validate server-side: a tampered/stale hidden
    form (key collision for op=price_change) is refused with nothing written."""
    data = {
        "op": "price_change", "site_id": SITE_ID, "product_code": PROD_CODE,
        "tenant_price": "182.0", "status": "tenanted",
        "valid_from": "2026-01-01",  # collides with the existing rule's key
        "reason": "tampered",
    }
    with FakeAirtable([_grid_rule("rec_old")]) as fa:
        with FakeAuthClient("admin") as client:
            r = client.post("/master/apply", data=data)
            assert r.status_code == 400, f"expected 400 refusal, got {r.status_code}"
            assert "already starts on that date" in r.text
    assert not fa.calls, f"a refused apply must write NOTHING, got {fa.calls}"
    assert fa.invalidations == 0


def test_route_cross_origin_post_rejected():
    """§3.3: _is_cross_origin on BOTH mutating POSTs — a foreign Origin is
    rejected before any parsing or writing."""
    with FakeAirtable([_grid_rule()]) as fa:
        with FakeAuthClient("admin") as client:
            for path in ("/master/preview", "/master/apply"):
                r = client.post(
                    path,
                    data={"op": "end_rule", "site_id": SITE_ID,
                          "product_code": PROD_CODE, "valid_from": "2026-01-01",
                          "valid_to": "2026-07-01", "reason": "csrf"},
                    headers={"Origin": "https://evil.example"},
                )
                assert r.status_code == 403, f"{path}: expected 403, got {r.status_code}"
    assert not fa.calls, "cross-origin POSTs must never reach a write"


def test_site_prefix_formula():
    """The scoped-read formula matches every rule_key starting with a site's
    "site|" prefix, OR-joins multiple sites, and bails (None -> full sweep) on an
    id that can't be safely embedded."""
    assert airtable_io._site_prefix_formula(["804"]) == 'FIND("804|",{rule_key})=1'
    assert airtable_io._site_prefix_formula(["804", "812"]) == (
        'OR(FIND("804|",{rule_key})=1,FIND("812|",{rule_key})=1)'
    )
    assert airtable_io._site_prefix_formula([]) is None
    assert airtable_io._site_prefix_formula(['80"4']) is None   # unsafe -> full read


def test_route_accept_overwrite_price_change_from_today():
    """The tenant-mismatch 'Set master to charged' button (overwrite=1) accepts
    the charged price as a FROM-TODAY price change: closes the live rule today,
    opens a successor at the charged price today (fb inherited). Never backdates."""
    today_iso = date.today().isoformat()
    rec = _grid_rule("rec_live", vf="2026-01-01")   # live £180 since January
    rec["fields"]["fb_price"] = 200.0
    with FakeAirtable([rec]) as fa:
        with FakeAuthClient("admin") as client:
            r = client.post("/accept-master-rule", data={
                "site_id": SITE_ID, "product_code": PROD_CODE,
                "tenant_price": "195.00", "overwrite": "1",
            })
            assert r.status_code == 200, r.text[:300]
            j = r.json()
            assert j["ok"] is True and "from today" in j["message"], j
            closes = [u for u in _updates(fa) if set(u["fields"]) == {"valid_to"}]
            assert [u["id"] for u in closes] == ["rec_live"]
            assert closes[0]["fields"]["valid_to"] == today_iso
            created = _creates(fa)
            assert len(created) == 1
            f = created[0]["fields"]
            assert f["tenant_price"] == 195.0 and f["valid_from"] == today_iso
            assert f["fb_price"] == 200.0, "fb inherited from the live rule"


def test_route_accept_overwrite_idempotent_same_price():
    """Overwriting to the SAME live price writes nothing and reports already-set."""
    rec = _grid_rule("rec_live", vf="2026-01-01")   # £180
    with FakeAirtable([rec]) as fa:
        with FakeAuthClient("admin") as client:
            r = client.post("/accept-master-rule", data={
                "site_id": SITE_ID, "product_code": PROD_CODE,
                "tenant_price": "180.00", "overwrite": "1",
            })
            assert r.status_code == 200
            j = r.json()
            assert j.get("ok") is True and j.get("already") is True, j
            assert not _updates(fa) and not _creates(fa), "no write when already at that price"


def test_findings_overwrite_button_render_and_guards():
    """Section 1 shows the overwrite button for admins only; a support row and a
    product charged at >1 price in the file get NO one-click button."""
    from summary import Summary, TenantSiteBlock, TenantRow, render_summary_html

    def _summary(rows):
        return Summary(
            file_name="wk.csv", line_count=len(rows), mismatch_count=len(rows),
            tenant_blocks=[TenantSiteBlock(site_id="001", site_name="Bar", rows=rows)],
            fb_blocks=[], missing_sites=[], products_not_on_master=[],
            tenant_price_missing=[], sites_in_sales_not_on_master=[], other_counts={},
        )

    single = _summary([TenantRow("P1", "Ale 9G", 180.0, 195.0, 2, 15.0, 30.0)])
    html_admin = render_summary_html(single, can_accept=True)
    assert 'data-overwrite="1"' in html_admin and "Set master to charged" in html_admin
    assert 'data-charged="195.00"' in html_admin and 'data-expected="180.00"' in html_admin
    html_view = render_summary_html(single, can_accept=False)
    assert "Set master to charged" not in html_view, "viewers get no accept button"

    # Support row -> no one-click overwrite (assert on the button marker, since
    # the section's explanatory intro paragraph also mentions the button name).
    supp = _summary([TenantRow("P1", "Ale 9G", 180.0, 195.0, 1, 15.0, 15.0, "support in effect")])
    assert 'data-overwrite="1"' not in render_summary_html(supp, can_accept=True)

    # Same product charged at two different prices -> ambiguous, route to editor.
    mixed = _summary([
        TenantRow("P1", "Ale 9G", 180.0, 195.0, 1, 15.0, 15.0),
        TenantRow("P1", "Ale 9G", 180.0, 200.0, 1, 20.0, 20.0),
    ])
    html_mixed = render_summary_html(mixed, can_accept=True)
    assert "use editor" in html_mixed and 'data-overwrite="1"' not in html_mixed


TESTS = [
    test_end_rule_closes_open_rule,
    test_end_rule_no_old_reason,
    test_end_rule_missing_key_raises,
    test_end_rule_already_ended_raises,
    test_end_rule_inverted_interval_raises,
    test_price_change_closes_prior_open_and_creates_successor,
    test_price_change_does_not_close_future_dated_sibling,
    test_price_change_same_key_patches_in_place,
    test_fix_in_place_updates_not_creates_and_prepends_reason,
    test_add_rule_new_key_creates_and_autocreates_site_product,
    test_apply_end_rule_dispatches_to_end_pricing_rule,
    test_validate_table,
    test_validate_never_forbids_overlap,
    test_preview_price_change_close_create_winner,
    test_preview_close_pass_skips_bounded_support_and_future_rule,
    test_backdated_price_change_behind_standing_open_rule_warns_with_window,
    test_compute_margin_math,
    test_mismatch_writer_is_idempotent_on_reupload,
    test_set_product_retro_updates_product_and_reflows_rules,
    test_render_master_pivot_shape_and_winner,
    test_render_master_pivot_cask_section_and_edit_mode,
    test_master_banner_escapes_source_filename,
    test_retro_pct_ge_one_is_blocked,
    test_retro_gbp_form_converts_to_fraction,
    test_preview_winner_overlapping_support_wins_today,
    test_preview_end_rule_delist_and_takeover,
    test_change_rule_key_normalises_codes,
    test_route_grid_viewer_ok_edit_admin_only,
    test_route_end_add_forms_and_404,
    test_route_preview_apply_roundtrip,
    test_route_cell_editor_amend_and_remove,
    test_route_site_rename,
    test_build_universal_increase_math,
    test_route_universal_increase_preview_apply,
    test_route_site_create_end_all_delete,
    test_route_product_settings_rename_end_all_delete,
    test_route_upload_master_updates_grid_immediately,
    test_route_export_master_from_snapshot,
    test_route_apply_revalidates_tampered_hidden_inputs,
    test_site_prefix_formula,
    test_route_accept_overwrite_price_change_from_today,
    test_route_accept_overwrite_idempotent_same_price,
    test_findings_overwrite_button_render_and_guards,
    test_route_cross_origin_post_rejected,
]


def main() -> int:
    failures = 0
    for t in TESTS:
        try:
            t()
        except AssertionError as e:
            failures += 1
            print(f"FAIL  {t.__name__}\n      {e}")
        except Exception as e:  # noqa: BLE001
            failures += 1
            print(f"ERROR {t.__name__}: {type(e).__name__}: {e}")
        else:
            print(f"ok    {t.__name__}")
    print(f"\n{len(TESTS) - failures}/{len(TESTS)} passed")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
