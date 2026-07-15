"""
Offline tests for the Tennents master workbook parser + §4 reconciliation.

Covers the README §4 spec end-to-end against a synthetic workbook:
expected rate from SKU_Master unless an exception overrides (raw-code keyed,
compound codes, resolved handling), ±£0.50/brl tolerance boundary, exact
retro-due arithmetic, line arithmetic, managed-site zero-retro handling,
bespoke (Gartocher-style) constructs, no-rate SKUs, unknown SKUs/accounts,
sign normalisation of the negative-convention monthly report, and barrelage
totals (incl. T.Lager via alt code).

No Airtable network access.

Run standalone (exit 0 = pass, 1 = fail):

    python test_tennents_master.py
"""
import os
import sys
import tempfile

# Windows consoles default to cp1252, which can't print Δ/±/£ in the labels.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook

from tennents import DeliveryLine, MonthlyReport, parse_monthly, reconcile
from tennents_master import SkuException, parse_master_workbook

PASS = True


def _check(label, cond, detail=""):
    global PASS
    PASS &= bool(cond)
    print(f"  [{'ok' if cond else 'FAIL'}] {label}{(' — ' + detail) if detail and not cond else ''}")


# ---------- fixture workbook ----------

def build_master_xlsx(path: str) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "README"
    ws.append(["Section", "Content"])
    ws.append(["4. Reconciliation logic", "tolerance ±£0.50/brl etc."])
    ws.append(["7. Version", "vTEST — synthetic fixture"])

    ws = wb.create_sheet("SKU_Master")
    ws.append(["SKU Code", "Alt Code", "Brand", "Product", "Container", "Brl per Unit",
               "ABV %", "WSP £/brl (post 2-Mar-26)", "Contract Base Discount £/brl",
               "On Contract Schedule?", "C&C / 3rd Party", "50% Hold £/brl (Mar-26)",
               "CURRENT CORRECT Total Discount £/brl", "Source", "Status / Notes"])
    #          code     alt       brand      product           cont  bpu     abv  wsp     base    onc  sup      hold   total   src        notes
    ws.append(["100100", "10010X", "Tennent's Lager", "T.Lager Keg", "11G", 0.3056, 4.0, 641.62, 302.00, "Y", "C&C", 12.33, 314.33, "Contract", "OK"])
    ws.append(["200200", None,     "Guinness",  "Guinness 50L", "50L", 0.3055, 4.1, 779.24, 249.00, "Y", "3rd party", 0, 249.00, "Contract", "OK"])
    ws.append(["200201", None,     "Guinness",  "Guinness 30L", "30L", 0.1833, 4.1, 779.24, 249.00, "Y", "3rd party", 0, 249.00, "Contract", "OK"])
    ws.append(["300300", "30030X", "Menabrea",  "Menabrea",     "50L", 0.3055, 4.8, 805.60, 316.52, "Y", "C&C",   15.48, 332.00, "Contract", "OK"])
    ws.append(["400400", None,     "NoRate",    "NoRate Keg",   "50L", 0.3055, 4.5, None,   None,   "N", "3rd party", 0, None,  "TBC",      "RATE TBC"])
    ws.append(["500500", None,     "BadArith",  "BadArith Keg", "50L", 0.3055, 4.0, 700.00, 300.00, "Y", "C&C",   10.00, 320.00, "Contract", "base+hold=310 ≠ 320"])

    ws = wb.create_sheet("Site_Master")
    ws.append(["Site", "Tennents Account", "Operating Model", "Discount Construct", "Notes"])
    ws.append(["STANDARD ARMS", 11110001, "Tenanted (TBC)", "Standard split (OID + retro)", None])
    ws.append(["MANAGED HOUSE", 11110002, "MANAGED (confirmed)", "ALL OFF-INVOICE, zero retro", None])
    ws.append(["BESPOKE BAR", 11110003, "Tenanted (confirmed)", "BESPOKE: flat £200/brl retro on all SKUs, OID as balancing figure", None])
    ws.append(["EXCEPTION INN", 11110004, "Tenanted (TBC)", "Standard split", None])
    ws.append(["GHOST TAVERN", 11110005, "Tenanted (TBC)", "Standard split", "never buys"])
    ws.append(["NEW SITE", "TBC", "Tenanted (TBC)", "Standard split", "account TBC"])
    ws.append(['ACTION: commentary row that must be skipped', None, None, None, None])

    ws = wb.create_sheet("Site_SKU_Exceptions")
    ws.append(["Site", "SKU", "Product", "Loaded Total Discount £/brl",
               "Correct Total Discount £/brl", "Direction / Who Bears", "£ Impact Jun-26", "Status"])
    # open exception: legacy low rate on the 30L Guinness code ONLY
    ws.append(["EXCEPTION INN", "200201", "Guinness 30L", 195.63, 249.00, "Under (split)", 9.78, "Raised 14-Jul-26"])
    # compound codes: both Menabrea containers share the override
    ws.append(["EXCEPTION INN", "300300/30030X", "Menabrea 50L/30L", 369.80, 332.00, "FB OVER", -18.49, "Monitoring — not raised"])
    # zero-loaded with a correct rate (credit pending)
    ws.append(["EXCEPTION INN", "100100", "T.Lager", 0, 314.33, "FB under", 96.06, "Correction requested"])
    # resolved via status text — must NOT override
    ws.append(["STANDARD ARMS", "200200", "Guinness 50L", 195.64, 249.00, "Under", 48.91, "RESOLVED 01-Jul-26 — rate corrected"])
    ws.append(["Amber = legend row that must be skipped", None, None, None, None, None, None, None])

    wb.save(path)


def build_monthly_xlsx(path: str) -> None:
    """Negative-convention Data sheet like the real Draught Pricing Report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ASM", "Customer Group", "Customer Name", "Brand", "SKU", "Year", "Month",
               "Kegs", "Barrels", "Invoice Price (per case/keg)",
               "Off invoice Discount per Brl", "Retro discount per Brl", "AOD per Brl",
               "Total Discount per Brl", "Retro Due", "Net Price per keg"])
    ws.append(["X", "FB", "STANDARD ARMS (11110001)", "TestLager", "T.LAGER KEG (100100)",
               2026, "2026/Jun", 2, 0.6112, 100.0, -100.0, -214.33, None, -314.33,
               -0.6112 * 214.33, 68.87])
    # a return line (kegs <= 0) — volume only, never price-checked
    ws.append(["X", "FB", "STANDARD ARMS (11110001)", "TestLager", "T.LAGER KEG (100100)",
               2026, "2026/Jun", -1, -0.3056, 100.0, -100.0, -214.33, None, -314.33,
               0.3056 * 214.33, 68.87])
    wb.save(path)


def L(account="11110001", customer="STANDARD ARMS", sku="100100", desc="T.Lager",
      kegs=1.0, barrels=0.3056, invoice=100.0, off=100.0, retro=214.33, aod=0.0,
      total=None, retro_due=None, net=None) -> DeliveryLine:
    if total is None:
        total = off + retro + aod
    if retro_due is None:
        retro_due = retro * barrels
    return DeliveryLine(account=account, customer_name=customer, sku_code=sku,
                        sku_desc=desc, kegs=kegs, barrels=barrels, invoice_price=invoice,
                        off_per_brl=off, retro_per_brl=retro, aod_per_brl=aod,
                        total_per_brl=total, retro_due=retro_due, net_price=net)


def R(*lines, excluded=(), period="2026-06") -> MonthlyReport:
    return MonthlyReport(lines=list(lines), excluded_lines=list(excluded),
                         period=period, sign_normalized=True)


def main() -> int:
    tmpdir = tempfile.mkdtemp(prefix="tennents_test_")
    master_path = os.path.join(tmpdir, "master.xlsx")
    monthly_path = os.path.join(tmpdir, "monthly.xlsx")
    build_master_xlsx(master_path)
    build_monthly_xlsx(monthly_path)

    print("workbook parsing:")
    m = parse_master_workbook(master_path, "master.xlsx")
    _check("version from README", m.version == "vTEST — synthetic fixture", m.version)
    _check("6 SKUs", len(m.skus) == 6, str(len(m.skus)))
    _check("6 sites (commentary row skipped)", len(m.sites) == 6, str(len(m.sites)))
    _check("4 exceptions (legend row skipped)", len(m.exceptions) == 4, str(len(m.exceptions)))
    _check("account normalised to str", m.sites[0].account == "11110001", m.sites[0].account)
    _check("managed flag", m.site_for_account("11110002").is_managed)
    _check("bespoke flat retro parsed", m.site_for_account("11110003").flat_retro_per_brl == 200.0)
    _check("no-rate SKU", m.find_sku("400400").correct_total_per_brl is None)
    _check("master arithmetic error found", [s.sku_code for s in m.arithmetic_errors()] == ["500500"])

    print("rate resolution:")
    _check("alt code hits SKU rate", m.resolve("11110001", "10010X").expected == 314.33)
    rb = m.resolve("11110004", "200201")
    _check("exception via raw code", rb.basis == "exception" and rb.expected == 195.63)
    _check("sibling container NOT overridden", m.resolve("11110004", "200200").basis == "sku_master")
    _check("compound exception code 1", m.resolve("11110004", "300300").basis == "exception")
    _check("compound exception code 2 (alt)", m.resolve("11110004", "30030X").basis == "exception")
    _check("resolved-status exception ignored", m.resolve("11110001", "200200").basis == "sku_master")
    _check("unknown sku", m.resolve("11110001", "ZZZ999").basis == "unknown_sku")
    _check("no-rate basis", m.resolve("11110001", "400400").basis == "no_rate")
    # resolved_flag=True retires an open exception (the Airtable checkbox path)
    ex_g = next(e for e in m.exceptions if e.sku_code_raw == "200201")
    ex_g.resolved_flag = True
    m.reindex()
    _check("resolved_flag retires override", m.resolve("11110004", "200201").basis == "sku_master")
    ex_g.resolved_flag = None
    m.reindex()

    print("monthly parsing (negative convention):")
    rep = parse_monthly(monthly_path)
    _check("period detected", rep.period == "2026-06", str(rep.period))
    _check("sign normalised", rep.sign_normalized)
    _check("1 checkable line", len(rep.lines) == 1)
    _check("1 excluded (return) line", len(rep.excluded_lines) == 1)
    ln = rep.lines[0]
    _check("discounts positive after normalise", ln.total_per_brl == 314.33 and ln.retro_per_brl == 214.33)
    _check("retro due normalised", abs(ln.retro_due - 0.6112 * 214.33) < 1e-9)

    print("reconcile — tolerance boundary (±£0.50/brl):")
    s = reconcile("t", m, R(L(total=314.33 - 0.50, retro=213.83)))
    _check("Δ = 0.50 exactly is OK", len(s.discount_mismatches) == 0)
    s = reconcile("t", m, R(L(total=314.33 - 0.51, retro=213.82)))
    _check("Δ = 0.51 flags", len(s.discount_mismatches) == 1)
    _check("positive Δ = short", s.discount_mismatches[0].delta_per_brl > 0)
    _check("delta_total = Δ × brl",
           abs(s.discount_mismatches[0].delta_total - 0.51 * 0.3056) < 1e-9)

    print("reconcile — exception handling:")
    # actual == loaded → pending, short vs correct tracked
    s = reconcile("t", m, R(L(account="11110004", customer="EXCEPTION INN",
                              sku="200201", off=100.0, retro=95.63, barrels=0.1833)))
    _check("known exception → pending, not mismatch",
           len(s.exception_pending) == 1 and len(s.discount_mismatches) == 0)
    _check("short vs correct = (249 − 195.63) × brl",
           abs(s.exception_pending[0].short_vs_correct - (249.00 - 195.63) * 0.1833) < 1e-6)
    # actual == correct → resolved signal
    s = reconcile("t", m, R(L(account="11110004", customer="EXCEPTION INN",
                              sku="200201", off=100.0, retro=149.0, barrels=0.1833)))
    _check("fix landed → exceptions_resolved", len(s.exceptions_resolved) == 1
           and len(s.exception_pending) == 0 and len(s.discount_mismatches) == 0)
    # matches neither → mismatch vs expected-current (loaded)
    s = reconcile("t", m, R(L(account="11110004", customer="EXCEPTION INN",
                              sku="200201", off=100.0, retro=120.0, barrels=0.1833)))
    _check("neither rate → mismatch vs loaded", len(s.discount_mismatches) == 1
           and s.discount_mismatches[0].basis == "exception expected-current"
           and abs(s.discount_mismatches[0].expected - 195.63) < 1e-9)
    # compound: both container codes hit the one override; aggregated by exception
    s = reconcile("t", m, R(
        L(account="11110004", customer="EXCEPTION INN", sku="300300", off=100.0, retro=269.80),
        L(account="11110004", customer="EXCEPTION INN", sku="30030X", off=100.0, retro=269.80),
    ))
    _check("compound codes → one pending row", len(s.exception_pending) == 1
           and s.exception_pending[0].sku_code == "300300/30030X"
           and abs(s.exception_pending[0].barrels - 2 * 0.3056) < 1e-9)
    _check("FB-over short is negative", s.exception_pending[0].short_vs_correct < 0)

    print("reconcile — retro exactness (§4):")
    s = reconcile("t", m, R(L()))
    _check("exact retro due passes", len(s.retro_arithmetic) == 0)
    s = reconcile("t", m, R(L(retro_due=214.33 * 0.3056 + 0.01)))
    _check("1p off flags", len(s.retro_arithmetic) == 1)
    s = reconcile("t", m, R(L(retro_due=None)))
    _check("absent Retro Due column → no check", len(s.retro_arithmetic) == 0)

    print("reconcile — line arithmetic:")
    s = reconcile("t", m, R(L(total=314.33, retro=214.33, off=100.0, aod=5.0)))
    _check("off+retro+AOD ≠ total flags", len(s.line_arithmetic) == 1)

    print("reconcile — managed sites:")
    s = reconcile("t", m, R(L(account="11110002", customer="MANAGED HOUSE",
                              off=314.33, retro=0.0, total=314.33, retro_due=0.0)))
    _check("zero retro + full OID is CORRECT (no findings)",
           not s.discount_mismatches and not s.managed_retro)
    s = reconcile("t", m, R(L(account="11110002", customer="MANAGED HOUSE",
                              off=100.0, retro=214.33, total=314.33)))
    _check("retro split at managed site → cash-timing flag",
           len(s.managed_retro) == 1 and not s.discount_mismatches)

    print("reconcile — bespoke construct (Gartocher-style):")
    s = reconcile("t", m, R(L(account="11110003", customer="BESPOKE BAR",
                              off=114.33, retro=200.0, total=314.33, retro_due=200.0 * 0.3056)))
    _check("flat £200/brl retro, total correct → clean",
           not s.discount_mismatches and not s.managed_retro and not s.retro_arithmetic)
    s = reconcile("t", m, R(L(account="11110003", customer="BESPOKE BAR",
                              off=100.0, retro=200.0, total=300.0, retro_due=200.0 * 0.3056)))
    _check("bespoke site: wrong TOTAL still flags", len(s.discount_mismatches) == 1)

    print("reconcile — unknowns:")
    s = reconcile("t", m, R(L(sku="400400", desc="NoRate Keg", total=100.0, off=100.0, retro=0.0, retro_due=0.0)))
    _check("no-rate SKU → no_rate row", len(s.no_rate) == 1 and not s.discount_mismatches)
    s = reconcile("t", m, R(L(sku="ZZZ999", desc="Mystery Keg")))
    _check("unknown SKU → not_on_master", len(s.not_on_master) == 1)
    s = reconcile("t", m, R(L(account="99999999", customer="MYSTERY BAR")))
    _check("unknown account → new_customers, no price checks",
           s.new_customers == [("99999999", "MYSTERY BAR")] and not s.discount_mismatches)

    print("reconcile — estate + volumes:")
    s = reconcile("t", m, R(L(), excluded=(L(kegs=-1.0, barrels=-0.3056),)))
    _check("returns net off barrelage", abs(s.barrels_total - 0.0) < 1e-9)
    s = reconcile("t", m, R(L(sku="10010X")))
    _check("T.Lager barrels via alt code", abs(s.tlager_barrels - 0.3056) < 1e-9)
    ghost = [name for _, name in s.sites_did_not_buy]
    _check("did-not-buy includes ghost + TBC sites",
           "GHOST TAVERN" in ghost and "NEW SITE" in ghost and "STANDARD ARMS" not in ghost)
    _check("master arithmetic surfaced", len(s.master_arithmetic) == 1
           and s.master_arithmetic[0].sku_code == "500500")

    print("render smoke:")
    from tennents import render_summary_html
    html = render_summary_html(reconcile("t", m, R(L(total=200.0, retro=100.0))))
    _check("summary renders", "Discount mismatches" in html and "vTEST" in html)

    print("PASS" if PASS else "FAIL")
    return 0 if PASS else 1


if __name__ == "__main__":
    sys.exit(main())
