"""
Offline tests for the team-facing Tennents price-file export.

Covers: the off-invoice/retro split by construct (standard uses the stored
per-site off-invoice; managed = all off-invoice; bespoke = flat retro), the
agreed rate coming from SKU_Master (not exceptions), no-rate SKUs shown as
RATE TBC, keg/pint maths (incl. multi-container keg-factor), a WSP change
flowing through to the keg price while the off-invoice holds, exception notes,
the Site_Prices parser round-trip, and that all three workbook builders emit
openable files. No Airtable network access.

    python test_tennents_price_export.py
"""
import io
import sys
import tempfile
import zipfile

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook, load_workbook

import tennents_price_export as tpe
from tennents_master import (
    SiteInfo,
    SitePrice,
    SkuException,
    SkuRate,
    TennentsMaster,
    keg_brl_factor,
    parse_master_workbook,
)

PASS = True


def _check(label, cond, detail=""):
    global PASS
    PASS &= bool(cond)
    print(f"  [{'ok' if cond else 'FAIL'}] {label}{(' — ' + detail) if detail and not cond else ''}")


def _near(a, b, tol=0.02):
    return a is not None and b is not None and abs(a - b) <= tol


# ---------- fixture ----------

def make_master() -> TennentsMaster:
    skus = [
        # 11G/22G multi-container, master brl_per_unit blank -> parse 11G
        SkuRate("090425", "09000X", "Tennent's Lager", "T.Lager 11G/22G", "11G / 22G",
                None, 4.0, 641.62, 302.00, True, "C&C", 12.33, 314.33),
        # 50L single, brl_per_unit present
        SkuRate("400889", "", "Jeffrey's", "Jeffrey's 50L", "50L",
                0.3055, 3.4, 538.39, 291.52, True, "C&C", 11.45, 302.97),
        # 50L/30L multi-container -> parse 50L
        SkuRate("GUI002", "GUI003", "Guinness", "Guinness 50L/30L", "50L / 30L",
                None, 4.1, 779.24, 249.00, True, "3rd party", 0.0, 249.00),
        # no agreed rate
        SkuRate("401211", "", "Tennent's", "T.Bavarian 50L", "50L",
                0.3055, 4.7, None, None, False, "C&C", 0.0, None),
    ]
    sites = [
        SiteInfo("11110001", "STANDARD ARMS", "Tenanted (TBC)", "Standard split"),
        SiteInfo("11110002", "MANAGED HOUSE", "MANAGED (confirmed)", "ALL OFF-INVOICE, zero retro"),
        SiteInfo("11110003", "BESPOKE BAR", "Tenanted (confirmed)",
                 "BESPOKE: flat £200/brl retro on all SKUs"),
    ]
    exceptions = [
        # open exception at STANDARD ARMS for Jeffrey's -> should surface as a note,
        # but the priced rate stays the AGREED 302.97 (not the loaded 291.54).
        SkuException("STANDARD ARMS", "11110001", "400889", "Jeffrey's 50L", 291.54, 302.97,
                     "FB under", 3.49, "open"),
    ]
    site_prices = [
        # STANDARD ARMS: T.Lager tenant gets £105.02/brl off-invoice; Jeffrey's full WSP (0)
        SitePrice("11110001", "STANDARD ARMS", "090425", "T.Lager", 105.02),
        SitePrice("11110001", "STANDARD ARMS", "400889", "Jeffrey's", 0.0),
        # A site-price given against the ALT code should still resolve to canonical
        SitePrice("11110001", "STANDARD ARMS", "GUI003", "Guinness", 60.00),
    ]
    return TennentsMaster("vTEST", "fixture", skus, sites, exceptions, site_prices)


def line(m, site_name, code):
    return tpe._price_line(m, tpe.find_site(m, site_name), m.find_sku(code))


def run():
    m = make_master()
    bpu_lager = keg_brl_factor(m.find_sku("090425"))   # 11G ≈ 0.3056
    bpu_gui = keg_brl_factor(m.find_sku("GUI002"))      # 50L ≈ 0.3055

    # ---- keg factor ----
    _check("11G keg factor ≈ 0.3056", _near(bpu_lager, 0.3056, 0.0005), f"{bpu_lager}")
    _check("50L/30L keg factor parses 50L ≈ 0.3055", _near(bpu_gui, 0.3055, 0.0005), f"{bpu_gui}")
    _check("brl_per_unit present is used", _near(keg_brl_factor(m.find_sku("400889")), 0.3055, 0.0005))

    # ---- standard site: stored off-invoice drives the split ----
    l = line(m, "STANDARD ARMS", "090425")
    _check("standard: agreed total 314.33", _near(l["total"], 314.33))
    _check("standard: off-invoice 105.02 from layer", _near(l["off"], 105.02))
    _check("standard: retro = total - off", _near(l["retro"], 314.33 - 105.02))
    _check("standard: net/brl = WSP - total", _near(l["net_brl"], 641.62 - 314.33))
    _check("standard: net/keg = (WSP-off)*bpu", _near(l["net_keg"], (641.62 - 105.02) * bpu_lager))
    _check("standard: net/pint = (WSP-off)/288", _near(l["net_pint"], (641.62 - 105.02) / 288.0, 0.001))

    # ---- off-invoice resolves via ALT code ----
    lg = line(m, "STANDARD ARMS", "GUI002")
    _check("alt-code site price resolves (GUI003 -> GUI002)", _near(lg["off"], 60.00))

    # ---- SKU with no stored site price -> off 0 (full WSP) ----
    lj = line(m, "STANDARD ARMS", "400889")
    _check("no stored off -> 0 (full WSP)", _near(lj["off"], 0.0))
    _check("full-WSP keg = WSP*bpu", _near(lj["net_keg"], 538.39 * 0.3055, 0.05))
    _check("open exception surfaces as a note", "Correction pending" in lj["note"], lj["note"])
    _check("priced rate is AGREED not loaded", _near(lj["total"], 302.97))

    # ---- "sold" = a live off-invoice DEAL (off > £0), NOT mere row-presence:
    # 090425 has a £105 off (sold); 400889 has a row but £0 off (no deal) ----
    _check("has_deal True for off>0 (090425)", line(m, "STANDARD ARMS", "090425")["has_deal"] is True)
    _check("has_deal False for £0-off row (400889)", line(m, "STANDARD ARMS", "400889")["has_deal"] is False)
    _check("has_deal True via alt code (GUI002)", line(m, "STANDARD ARMS", "GUI002")["has_deal"] is True)
    _check("has_deal False for unpriced SKU (401211)", line(m, "STANDARD ARMS", "401211")["has_deal"] is False)

    # ---- managed site: all off-invoice, zero retro ----
    lm = line(m, "MANAGED HOUSE", "GUI002")
    _check("managed: off = total", _near(lm["off"], 249.00))
    _check("managed: retro = 0", _near(lm["retro"], 0.0))
    _check("managed: net/keg = (WSP-total)*bpu", _near(lm["net_keg"], (779.24 - 249.00) * bpu_gui))
    _check("managed: note flags it", "Managed" in lm["note"], lm["note"])

    # ---- bespoke: master does NOT model the split (README §4) — use stored
    # off (0 here), tenant on full WSP; construct surfaced as a note only ----
    lb = line(m, "BESPOKE BAR", "GUI002")
    _check("bespoke: off = stored (0, no split invented)", _near(lb["off"], 0.0))
    _check("bespoke: retro = total", _near(lb["retro"], 249.00))
    _check("bespoke: construct note", "Bespoke" in lb["note"], lb["note"])

    # ---- no agreed rate -> RATE TBC ----
    lt = line(m, "STANDARD ARMS", "401211")
    _check("no-rate: total None", lt["total"] is None)
    _check("no-rate: net/keg None", lt["net_keg"] is None)
    _check("no-rate: RATE TBC note", "RATE TBC" in lt["note"], lt["note"])

    # ---- WSP change flows through; off-invoice holds ----
    m.find_sku("090425").wsp_per_brl = 671.62      # +30 PINC
    m.reindex()
    l2 = line(m, "STANDARD ARMS", "090425")
    _check("PINC: off-invoice unchanged", _near(l2["off"], 105.02))
    _check("PINC: net/keg rose by 30*bpu", _near(l2["net_keg"], (671.62 - 105.02) * bpu_lager))
    _check("PINC: net/brl uses new WSP - total", _near(l2["net_brl"], 671.62 - 314.33))

    # ---- workbook builders emit openable files ----
    mb = tpe.build_master_workbook_bytes(m)
    wb = load_workbook(io.BytesIO(mb))
    _check("master workbook: a tab per site", len(wb.sheetnames) == 3, str(wb.sheetnames))
    _check("master workbook: tab named for site", "STANDARD ARMS" in wb.sheetnames)

    # Two purpose-built sections: "Currently sold" (products with an off-invoice
    # deal) then "Substitution pricing" (the rest), each grouped by beer type.
    ws = wb["STANDARD ARMS"]
    colA = [str(ws.cell(row=rr, column=1).value or "") for rr in range(1, ws.max_row + 1)]
    sold_title = next((i for i, v in enumerate(colA) if v.startswith("Currently sold")), None)
    sub_title = next((i for i, v in enumerate(colA) if v.startswith("Substitution pricing")), None)
    _check("two sections, sold before substitution",
           sold_title is not None and sub_title is not None and sold_title < sub_title,
           f"sold@{sold_title} sub@{sub_title}")

    hdr_rows = [rr for rr in range(1, ws.max_row + 1) if ws.cell(row=rr, column=1).value == "SKU Code"]
    _check("two section header rows", len(hdr_rows) == 2, str(hdr_rows))
    sold_h, sub_h = (hdr_rows + [None, None])[:2]

    def _row_for(code):
        for rr in range(1, ws.max_row + 1):
            if ws.cell(row=rr, column=1).value == code:
                return rr
        return None

    r_sold = _row_for("090425")   # £105 off -> a deal -> Currently sold
    r_sub = _row_for("400889")    # £0 off -> no deal -> Substitution candidate
    if sold_h and sub_h and r_sold and r_sub:
        _check("off>0 product in Currently-sold", sold_h < r_sold < sub_h, f"{sold_h}<{r_sold}<{sub_h}")
        _check("£0-off product in Substitution", r_sub > sub_h, f"{r_sub}>{sub_h}")

        # SOLD columns: Off-Invoice=F(6) input; Net Keg=G(7)=(D-F)*bpu; FB Retro=I(9)=E-F
        _check("sold header: Off-Invoice at F", ws.cell(row=sold_h, column=6).value == "Tenant Off-Invoice £/Brl")
        _check("sold header: FB Retro at I", ws.cell(row=sold_h, column=9).value == "FB Retro £/Brl")
        _check("sold: off-invoice is a numeric input", isinstance(ws.cell(row=r_sold, column=6).value, (int, float)))
        _check("sold: Net Keg = (D-F)*bpu live",
               str(ws.cell(row=r_sold, column=7).value).startswith(f"=(D{r_sold}-F{r_sold})*"),
               str(ws.cell(row=r_sold, column=7).value))
        _check("sold: FB Retro = E-F live", ws.cell(row=r_sold, column=9).value == f"=E{r_sold}-F{r_sold}",
               str(ws.cell(row=r_sold, column=9).value))

        # SUBSTITUTION columns: FB cost=F(6); per RPB an off-invoice AND a net keg
        _check("sub header: FB Net Cost at F", ws.cell(row=sub_h, column=6).value == "FB Net Cost £/Keg")
        _check("sub header: Off-Inv @£150 at G", ws.cell(row=sub_h, column=7).value == "Off-Inv £/Brl @£150")
        _check("sub header: Net @£150 at H", ws.cell(row=sub_h, column=8).value == "Net £/Keg @£150")
        _check("sub: FB net cost = (D-E)*bpu",
               str(ws.cell(row=r_sub, column=6).value).startswith(f"=(D{r_sub}-E{r_sub})*"),
               str(ws.cell(row=r_sub, column=6).value))
        _check("sub: off-invoice @£150 = Total - 150", ws.cell(row=r_sub, column=7).value == f"=E{r_sub}-150",
               str(ws.cell(row=r_sub, column=7).value))
        _check("sub: net @£150 = (D-E+150)*bpu",
               str(ws.cell(row=r_sub, column=8).value).startswith(f"=(D{r_sub}-E{r_sub}+150)*"),
               str(ws.cell(row=r_sub, column=8).value))

        _check("Notes at col 13 in both sections",
               ws.cell(row=sold_h, column=13).value == "Notes" and ws.cell(row=sub_h, column=13).value == "Notes")

        # beer-type grouping kept within each section
        def _band_above(rr):
            cats = set(tpe._CATEGORY_ORDER) | {tpe._OTHER}
            for k in range(rr - 1, 0, -1):
                v = ws.cell(row=k, column=1).value
                if v in cats:
                    return v
                if v == "SKU Code":
                    return None
            return None
        _check("sold product grouped by beer type", _band_above(r_sold) == "Standard Lager", str(_band_above(r_sold)))
        _check("sub product grouped by beer type", _band_above(r_sub) == "Standard Lager", str(_band_above(r_sub)))

    # TBC product (401211) has no deal -> substitution section, unpriced -> blank
    rtbc = _row_for("401211")
    _check("TBC product in substitution section", rtbc is not None and sub_h and rtbc > sub_h, str(rtbc))
    if rtbc:
        _check("TBC row: FB cost blank", ws.cell(row=rtbc, column=6).value in (None, ""),
               str(ws.cell(row=rtbc, column=6).value))
        _check("TBC row: net @£150 blank", ws.cell(row=rtbc, column=8).value in (None, ""),
               str(ws.cell(row=rtbc, column=8).value))

    # managed site: no split, a single section
    wsm = wb["MANAGED HOUSE"]
    mh_hdrs = [rr for rr in range(1, wsm.max_row + 1) if wsm.cell(row=rr, column=1).value == "SKU Code"]
    _check("managed site: single section (no split)", len(mh_hdrs) == 1, str(mh_hdrs))

    # page break so the substitution list prints as a second page
    try:
        n_breaks = len(ws.row_breaks.brk)
    except Exception:
        n_breaks = -1
    _check("page break set between sections", n_breaks >= 1, str(n_breaks))

    sb = tpe.build_single_site_bytes(m, tpe.find_site(m, "MANAGED HOUSE"))
    wb1 = load_workbook(io.BytesIO(sb))
    _check("single-site workbook: one tab", len(wb1.sheetnames) == 1)

    zb = tpe.build_all_sites_zip_bytes(m)
    zf = zipfile.ZipFile(io.BytesIO(zb))
    _check("zip: one file per site", len(zf.namelist()) == 3, str(zf.namelist()))
    _check("zip: site name in filename", any("STANDARD ARMS" in n for n in zf.namelist()))

    _check("site choices listed", len(tpe.list_site_choices(m)) == 3)
    _check("find_site by account", tpe.find_site(m, "11110002").site_name == "MANAGED HOUSE")
    _check("find_site by name", tpe.find_site(m, "bespoke bar").account == "11110003")

    # ---- Site_Prices parser round-trip ----
    _roundtrip_parse()

    print("\n" + ("ALL PASS" if PASS else "FAILURES ABOVE"))
    return 0 if PASS else 1


def _roundtrip_parse():
    """Write a minimal master workbook WITH a Site_Prices sheet and parse it."""
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.append(["Section", "Content"])
    ws.append(["7. Version", "vRT"])

    ws = wb.create_sheet("SKU_Master")
    ws.append(["SKU Code", "Alt Code", "Brand", "Product", "Container", "Brl per Unit",
               "ABV %", "WSP £/brl", "Contract Base Discount £/brl", "On Contract Schedule?",
               "C&C / 3rd Party", "50% Hold £/brl", "CURRENT CORRECT Total Discount £/brl",
               "Source", "Status / Notes"])
    ws.append(["090425", "09000X", "T.Lager", "T.Lager", "11G", 0.3056, 4.0, 641.62,
               302.0, "Y", "C&C", 12.33, 314.33, "src", "ok"])

    ws = wb.create_sheet("Site_Master")
    ws.append(["Site", "Tennents Account", "Operating Model", "Discount Construct", "Notes"])
    ws.append(["STANDARD ARMS", 11110001, "Tenanted", "Standard split", None])

    ws = wb.create_sheet("Site_SKU_Exceptions")
    ws.append(["Site", "SKU", "Product", "Loaded Total Discount £/brl",
               "Correct Total Discount £/brl", "Direction / Who Bears", "£ Impact", "Status"])

    ws = wb.create_sheet("Site_Prices")
    ws.append(["Site", "Tennents Account", "SKU Code", "Product",
               "Off-Invoice Discount £/Brl", "Notes"])
    # account left blank on purpose -> resolved from Site_Master by name
    ws.append(["STANDARD ARMS", None, "090425", "T.Lager", 105.02, ""])

    path = tempfile.mktemp(suffix=".xlsx")
    wb.save(path)
    m = parse_master_workbook(path, "rt")
    _check("parser: Site_Prices row loaded", len(m.site_prices) == 1, str(len(m.site_prices)))
    _check("parser: site_prices_present True when sheet exists", m.site_prices_present is True)
    _check("parser: account resolved from name", m.site_prices[0].account == "11110001",
           m.site_prices[0].account)
    _check("parser: off_invoice resolves", _near(m.off_invoice("11110001", "090425"), 105.02))
    _check("parser: off_invoice via alt code", _near(m.off_invoice("11110001", "09000X"), 105.02))

    # a workbook WITHOUT Site_Prices still parses (backward compatible)
    del m
    wb2 = load_workbook(path)
    del wb2["Site_Prices"]
    path2 = tempfile.mktemp(suffix=".xlsx")
    wb2.save(path2)
    m2 = parse_master_workbook(path2, "rt2")
    _check("parser: absent Site_Prices is fine", m2.site_prices == [])
    _check("parser: site_prices_present False when absent (preserve signal)",
           m2.site_prices_present is False)
    _check("parser: off_invoice defaults 0 with no layer", m2.off_invoice("11110001", "090425") == 0.0)


if __name__ == "__main__":
    raise SystemExit(run())
